'''
Copyright 2023

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

from app import app, db, celery
from app.models import *
from app.functions.globals import retryTime, list_all, chunker, batch_crops, rDets, randomString, stringify_timestamp, getChildList, \
    resetImageDownloadStatus, resetVideoDownloadStatus, deleteFile, getChildNameList
import GLOBALS
from sqlalchemy.sql import alias, func, or_, and_, distinct, case
import re
import math
import ast
import boto3
from config import Config
import os
from multiprocessing.pool import ThreadPool as Pool
import tempfile
import traceback
import piexif
from iptcinfo3 import IPTCInfo
import shutil
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font
import pandas as pd
import json
import io
from celery.result import allow_join_result
import redis
from datetime import datetime, timedelta
import numpy as np
from app.functions.imports import s3traverse
from app.functions.permissions import surveyPermissionsSQ
import zipfile
import csv
import tracemalloc

def translate(labels, dictionary):
    '''
    Helper function for prepareComparison. Translates label IDs to index numbers.

        Parameters:
            labels (list): list of label IDs for translation
            dictionary (dict): dictionary that translates IDs to index numbers

        Returns:
            newLabels (list): list of label indexes
    '''
    
    newLabels = []
    for label in labels:
        if str(label) == 'nan':
            pass
        else:
            newLabel = dictionary[str(math.trunc(label))]
            if newLabel not in newLabels:
                newLabels.append(newLabel)

    if newLabels == []:
        newLabels = [dictionary['None']]

    return newLabels

def compareLabels(user_id,image_id,labels1,labels2,nothing_label1,nothing_label2,ground_truth,detection,detection_source):
    '''
    Helper function for prepareComparison. Compares the labels on an image-by-image basis, and populates the comparison global variables.

        Parameters:
            user_id (int): The user requesting th comparison
            image_id (int): The image for which the comparison is performed
            labels1 (list): List of label indices for the image from task 1
            labels2 (list): List of label indices for the image from task 2
            nothing_label1 (str): Index of the nothing label category of task 1
            nothing_label2 (str): Index of the nothing label category of task 2
            ground_truth (int): The ground truth task in the comparison
            detection (floar): Detection confidence threshold
            detection_source (list): The source of the detections
    '''
    
    matches = []
    for label in labels1:
        if label in labels2:
            GLOBALS.confusions[label][label].append(image_id)
            matches.append(label)

    for match in matches:
        labels1.remove(match)
        labels2.remove(match)

    # MegaDetector misses
    if ground_truth == 1:
        if (nothing_label2 in labels2) and (nothing_label1 not in labels1) and (detection<=Config.DETECTOR_THRESHOLDS[detection_source[0]]):
            GLOBALS.megaDetectorMisses['image_ids'].append(image_id)
            GLOBALS.megaDetectorMisses['count'] += len(labels1)
    else:
        if (nothing_label1 in labels1) and (nothing_label2 not in labels2) and (detection<=Config.DETECTOR_THRESHOLDS[detection_source[0]]):
            GLOBALS.megaDetectorMisses['image_ids'].append(image_id)
            GLOBALS.megaDetectorMisses['count'] += len(labels2)

    if (len(labels1) != 0) or ((len(labels2) != 0)):
        if (len(labels1) > 1) or ((len(labels2) > 1)):
            if labels1 == [nothing_label1]:
                for label in labels2:
                    GLOBALS.confusions[nothing_label1][label].append(image_id)
            elif labels2 == [nothing_label2]:
                for label in labels1:
                    GLOBALS.confusions[label][nothing_label2].append(image_id)
            else:
                GLOBALS.confusions['multi'].append(image_id)
        else:
            if len(labels1) == 0:
                GLOBALS.confusions[nothing_label1][labels2[0]].append(image_id)
            elif len(labels2) == 0:
                GLOBALS.confusions[labels1[0]][nothing_label2].append(image_id)
            else:
                GLOBALS.confusions[labels1[0]][labels2[0]].append(image_id)

    return True

def findEmptyClustered(user_id,image_id,cluster_labels,image_labels,nothing_label):
    '''
    Helper function for prepareComparison. Finds empty images that have been clustered into non-empty clusters and populates comparison globals.

        Parameters:
            user_id (int): The user for which the comparison is being prepared
            image_id (int): The image being checked
            cluster_labels (list): The cluster labels
            image_labels (list): The label indices for the image
            nothing_label (str): Index of the nothing label category
    '''
    
    if (len(cluster_labels)>1) and (nothing_label in cluster_labels) and (nothing_label in image_labels):
        GLOBALS.emptyClustered.append(image_id)
    return True

@celery.task(bind=True,max_retries=1,ignore_result=True)
def prepareComparison(self,translations,groundTruth,task_id1,task_id2,user_id):
    '''
    Prepares the requested task comparison for a survey by populating the comparison globals.

        Parameters:
            translations (str): Raw translation string from user
            groundTruth (int): ID of task being used as the ground truth
            task_id1 (int): First task for comparison purposes
            task_id2 (int): Second task for comparison purposes
            user_id (int): User that requested the comparison
    '''
    
    try:
        app.logger.info('Preparing comparison.')
        task_id1 = int(task_id1)
        task_id2 = int(task_id2)
        task1 = db.session.query(Task).get(task_id1)
        survey_id = task1.survey_id
        translations = translations.replace('*****', '/')
        translations = ast.literal_eval(translations)

        ground_truths = {}
        ground_truths['ground'] = int(groundTruth)
        ground_truths['task1'] = int(task_id1)
        ground_truths['task2'] = int(task_id2)

        if int(groundTruth) == int(task_id1):
            ground_truths['other'] = int(task_id2)
            ground_truth = 1
        else:
            ground_truths['other'] = int(task_id1)
            ground_truth = 2

        task1Translation = {}
        task2Translation = {}
        newLabels = {}
        index = 1
        for key in translations:
            newLabels[str(index)] = key
            for labelID in translations[key][str(task_id1)]:
                task1Translation[str(labelID)] = str(index)
            for labelID in translations[key][str(task_id2)]:
                task2Translation[str(labelID)] = str(index)
            index += 1

        GLOBALS.redisClient.set('comparisonLabels_'+str(user_id), json.dumps(newLabels))
        
        task1Translation['None'] = task1Translation[str(GLOBALS.nothing_id)]
        task2Translation['None'] = task2Translation[str(GLOBALS.nothing_id)]

        ground_truths['nothing1'] = task1Translation[str(GLOBALS.nothing_id)]
        ground_truths['nothing2'] = task2Translation[str(GLOBALS.nothing_id)]

        # detection-label based
        sq1 = db.session.query(Image.id.label('image_id1'),Label.id.label('label_id1'))\
                        .join(Detection)\
                        .join(Labelgroup)\
                        .join(Label, Labelgroup.labels)\
                        .filter(Labelgroup.task_id==task_id1)\
                        .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                        .filter(Detection.static==False)\
                        .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                        .subquery()
        sq2 = db.session.query(Image.id.label('image_id2'),Label.id.label('label_id2'))\
                        .join(Detection)\
                        .join(Labelgroup)\
                        .join(Label, Labelgroup.labels)\
                        .filter(Labelgroup.task_id==task_id2)\
                        .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                        .filter(Detection.static==False)\
                        .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                        .subquery()

        if ground_truth == 1:
            df = pd.read_sql(db.session.query( \
                            Image.id.label('image_id'), \
                            sq1.c.label_id1.label('label_id1'), \
                            sq2.c.label_id2.label('label_id2'),\
                            Cluster.id.label('cluster_id'), \
                            Detection.score.label('detections'),
                            Detection.source.label('detection_source')) \
                            .outerjoin(sq1, sq1.c.image_id1==Image.id) \
                            .outerjoin(sq2, sq2.c.image_id2==Image.id) \
                            .outerjoin(Detection, Detection.image_id==Image.id) \
                            .join(Camera,Image.camera_id==Camera.id) \
                            .join(Trapgroup,Camera.trapgroup_id==Trapgroup.id) \
                            .join(Cluster, Image.clusters) \
                            .filter(Trapgroup.survey_id==survey_id) \
                            .filter(Cluster.task_id==task_id2) \
                            .statement,db.session.bind)
        else:
            df = pd.read_sql(db.session.query( \
                            Image.id.label('image_id'), \
                            sq1.c.label_id1.label('label_id1'), \
                            sq2.c.label_id2.label('label_id2'),\
                            Cluster.id.label('cluster_id'), \
                            Detection.score.label('detections'),
                            Detection.source.label('detection_source')) \
                            .outerjoin(sq1, sq1.c.image_id1==Image.id) \
                            .outerjoin(sq2, sq2.c.image_id2==Image.id) \
                            .outerjoin(Detection, Detection.image_id==Image.id) \
                            .join(Camera,Image.camera_id==Camera.id) \
                            .join(Trapgroup,Camera.trapgroup_id==Trapgroup.id) \
                            .join(Cluster, Image.clusters) \
                            .filter(Trapgroup.survey_id==survey_id) \
                            .filter(Cluster.task_id==task_id1) \
                            .statement,db.session.bind)

        df = df.groupby('image_id').agg(lambda x: tuple(set(x))).reset_index()

        df['detection'] = df['detections'].apply(max)
        df['cluster_detection'] = df.groupby('cluster_id')['detection'].transform(max)
        del df['detections']
        del df['detection']

        df['task1_labels'] = df.apply(lambda x: translate(x.label_id1, task1Translation), axis=1)
        df['task2_labels'] = df.apply(lambda x: translate(x.label_id2, task2Translation), axis=1)

        del df['label_id1']
        del df['label_id2']

        GLOBALS.confusions = {}
        GLOBALS.confusions['multi'] = []
        for i in range(1,index):
            GLOBALS.confusions[str(i)] = {}
            for n in range(1,index):
                GLOBALS.confusions[str(i)][str(n)] = []

        GLOBALS.megaDetectorMisses = {}
        GLOBALS.megaDetectorMisses['image_ids'] = []
        GLOBALS.megaDetectorMisses['count'] = 0

        nothing_label1 = task1Translation['None']
        nothing_label2 = task2Translation['None']
        df.apply(lambda x: compareLabels(user_id,x.image_id,x.task1_labels,x.task2_labels,nothing_label1,nothing_label2,ground_truth,x.cluster_detection,x.detection_source), axis=1)

        # Find error souces:
        # empty-clustered
        GLOBALS.emptyClustered = []
        if ground_truth == 1:
            df['cluster_labels1'] = df.groupby('cluster_id')['task1_labels'].transform(sum)
            df['cluster_labels'] = df.apply(lambda x: set(x.cluster_labels1), axis=1)
            del df['cluster_labels1']
            df.apply(lambda x: findEmptyClustered(user_id,x.image_id,x.cluster_labels,x.task1_labels,nothing_label1), axis=1)
        else:
            df['cluster_labels1'] = df.groupby('cluster_id')['task2_labels'].transform(sum)
            df['cluster_labels'] = df.apply(lambda x: set(x.cluster_labels1), axis=1)
            del df['cluster_labels1']
            df.apply(lambda x: findEmptyClustered(user_id,x.image_id,x.cluster_labels,x.task2_labels,nothing_label2), axis=1)

        GLOBALS.redisClient.set('megaDetectorMisses_'+str(user_id), json.dumps(GLOBALS.megaDetectorMisses))
        GLOBALS.redisClient.set('emptyClustered_'+str(user_id), json.dumps(GLOBALS.emptyClustered))
        GLOBALS.redisClient.set('ground_truths_'+str(user_id), json.dumps(ground_truths))
        GLOBALS.redisClient.set('confusions_'+str(user_id), json.dumps(GLOBALS.confusions))

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        self.retry(exc=exc, countdown= retryTime(self.request.retries))

    finally:
        db.session.remove()

    return True

def generate_url(rootUrl,level_type,x_level):
    ''' Helper function for create_task_dataframe that collapses the urls generated for videos based on the argument.'''
    return rootUrl + level_type + '&id=' + str(x_level)

def create_task_dataframe(task_id,selectedLevel,include,exclude,trapgroup_id,startDate,endDate,required_columns):
    '''
    Returns an all-encompassing dataframe for a task, subject to the parameter selections.

        Parameters:
            task_id (int): Task for which dataframe has been requested
            selectedLevel (str): The data for each for in the csv
            include (list): The label ids to include
            exclude (list): The label ids to exclude
            trapgroup_id (int): The trapgroup id to filter on
            startDate (datetime): The start date to filter on
            endDate (datetime): The end date to filter on

        Returns:
            df (pd.dataframe): task dataframe
    '''

    task = db.session.query(Task).get(task_id)
    
    query = db.session.query( \
                Image.id.label('Image ID'),\
                func.concat(
                    func.substr(Camera.path, func.locate('/', Camera.path, 1) + 1), # Skip first folder
                    '/',
                    Image.filename
                ).label('File'), \
                Image.corrected_timestamp.label('Timestamp'), \
                Image.timestamp.label('Original Timestamp'), \
                Detection.id.label('Sighting ID'), \
                Detection.left.label('Left'), \
                Detection.right.label('Right'), \
                Detection.top.label('Top'), \
                Detection.bottom.label('Bottom'), \
                Detection.score.label('Score'), \
                Cluster.notes.label('Notes'), \
                Cluster.id.label('Cluster ID'), \
                Label.description.label('Species'), \
                Tag.description.label('Informational Tags'), \
                Cameragroup.id.label('Camera ID'), \
                (Trapgroup.tag + '-' + Cameragroup.name).label('Camera Name'), \
                Trapgroup.id.label('Site ID'), \
                Trapgroup.tag.label('Site Name'), \
                Trapgroup.latitude.label('Latitude'), \
                Trapgroup.longitude.label('Longitude'), \
                Trapgroup.altitude.label('Altitude')) \
                .join(Image,Cluster.images) \
                .join(Detection,Detection.image_id==Image.id) \
                .join(Labelgroup,Labelgroup.detection_id==Detection.id) \
                .join(Label,Labelgroup.labels,isouter=True) \
                .join(Tag,Labelgroup.tags,isouter=True) \
                .join(Camera,Image.camera_id==Camera.id) \
                .join(Cameragroup,Camera.cameragroup_id==Cameragroup.id) \
                .join(Trapgroup,Camera.trapgroup_id==Trapgroup.id) \
                .filter(Cluster.task_id==task_id) \
                .filter(Labelgroup.task_id==task_id) \
                .filter(Detection.static==False) \
                .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS)) \
                .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))

    sq = db.session.query(Image)\
                .join(Detection)\
                .join(Camera)\
                .join(Trapgroup)\
                .filter(Trapgroup.survey_id==task.survey_id)\
                .filter(Detection.static==False)\
                .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))

    if len(include) != 0:
        query = query.filter(Label.id.in_(include))
        sq = sq.filter(Label.id.in_(include))

    if len(exclude) != 0:
        query = query.filter(~Label.id.in_(exclude))
        sq = sq.filter(~Label.id.in_(exclude))

    if GLOBALS.nothing_id in exclude:
        query = query.filter(Labelgroup.labels.any())
        sq = sq.filter(Labelgroup.labels.any())

    if trapgroup_id:
        query = query.filter(Trapgroup.id==trapgroup_id)
        sq = sq.filter(Trapgroup.id==trapgroup_id)

    if startDate:
        query = query.filter(Image.corrected_timestamp>=startDate)
        sq = sq.filter(Image.corrected_timestamp>=startDate)

    if endDate:
        query = query.filter(Image.corrected_timestamp<=endDate)
        sq = sq.filter(Image.corrected_timestamp<=endDate)

    df = pd.read_sql(query.statement,db.session.bind)

    if (len(include) == 0) and (GLOBALS.nothing_id not in exclude):
        sq = sq.subquery()

        #This includes all the images with no detections
        query = db.session.query( \
                        Image.id.label('Image ID'),\
                        func.concat(
                            func.substr(Camera.path, func.locate('/', Camera.path, 1) + 1), # Skip first folder
                            '/',
                            Image.filename
                        ).label('File'), \
                        Image.corrected_timestamp.label('Timestamp'), \
                        Image.timestamp.label('Original Timestamp'), \
                        Detection.id.label('Sighting ID'), \
                        Detection.left.label('Left'), \
                        Detection.right.label('Right'), \
                        Detection.top.label('Top'), \
                        Detection.bottom.label('Bottom'), \
                        Detection.score.label('Score'), \
                        Cluster.id.label('Cluster ID'), \
                        Cameragroup.id.label('Camera ID'), \
                        (Trapgroup.tag + '-' + Cameragroup.name).label('Camera Name'), \
                        Trapgroup.id.label('Site ID'), \
                        Trapgroup.tag.label('Site Name'), \
                        Trapgroup.latitude.label('Latitude'), \
                        Trapgroup.longitude.label('Longitude'), \
                        Trapgroup.altitude.label('Altitude')) \
                        .join(Image,Cluster.images) \
                        .join(Detection,Detection.image_id==Image.id) \
                        .join(Camera,Image.camera_id==Camera.id) \
                        .join(Cameragroup,Camera.cameragroup_id==Cameragroup.id) \
                        .join(Trapgroup,Camera.trapgroup_id==Trapgroup.id) \
                        .outerjoin(sq,sq.c.id==Image.id)\
                        .filter(Cluster.task_id==task_id)\
                        .filter(sq.c.id==None)

        if trapgroup_id: query = query.filter(Trapgroup.id==trapgroup_id)
        if startDate: query = query.filter(Image.corrected_timestamp>=startDate)
        if endDate: query = query.filter(Image.corrected_timestamp<=endDate)                     

        df2 = pd.read_sql(query.statement,db.session.bind)

        df2['Species'] = 'None'
        df2['Informational Tags'] = 'None'
        df = pd.concat([df,df2]).reset_index()
        df2 = None

    # Add individuals (they don't want to outer join)
    if 'Individuals' in required_columns:
        query  = db.session.query( \
                    Detection.id.label('Sighting ID'), \
                    Individual.name.label('Individuals'),\
                    Individual.species.label('Species')) \
                    .join(Image)\
                    .join(Camera)\
                    .join(Trapgroup)\
                    .join(Individual,Detection.individuals) \
                    .filter(Trapgroup.survey==task.survey)\
                    .filter(Individual.tasks.contains(task)) \
                    .filter(Individual.active==True)\
                    .filter(Detection.static==False) \
                    .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS)) \
                    .filter(~Detection.status.in_(['deleted','hidden']))

        if include: query = query.filter(Individual.species.in_([r[0] for r in db.session.query(Label.description).filter(Label.id.in_(include)).distinct().all()]))
        if exclude: query = query.filter(~Individual.species.in_([r[0] for r in db.session.query(Label.description).filter(Label.id.in_(exclude)).distinct().all()]))
        if trapgroup_id: query = query.filter(Trapgroup.id==trapgroup_id)
        if startDate: query = query.filter(Image.corrected_timestamp>=startDate)
        if endDate: query = query.filter(Image.corrected_timestamp<=endDate)   

        df3 = pd.read_sql(query.statement,db.session.bind)
        df = pd.merge(df, df3, on=['Sighting ID','Species'], how='outer')
        df3 = None

    # Cast to more efficient data types
    types = {
        'Left': 'float32',
        'Right': 'float32',
        'Top': 'float32',
        'Bottom': 'float32',
        'Score': 'float32',
        'Camera ID': 'int32',
        'Site ID': 'int32',
        'Latitude': 'float32',
        'Longitude': 'float32',
        'Altitude': 'float32',
        'Site Name': 'category',
        'Camera Name': 'category'
    } # Image, Detection & Cluster ID we leave as int64 (they are large)
    df = df.astype(types)

    #If df is empty, return it
    if len(df) == 0: return df

    # add video paths if needed
    if 'Video Path' in required_columns:
        video_query = db.session.query(
                                    Image.id.label('Image ID'),
                                    func.concat(
                                        func.substring_index(
                                            func.substr(Camera.path, func.locate('/', Camera.path, 1) + 1),"_video_images_", 1),
                                        Video.filename
                                    ).label('Video Path')
                                )\
                                .join(Camera,Image.camera_id==Camera.id)\
                                .join(Video,Camera.videos)
        
        if trapgroup_id:
            video_query = video_query.filter(Camera.trapgroup_id==trapgroup_id)
        else:
            video_query = video_query.join(Trapgroup).filter(Trapgroup.survey_id==task.survey_id)

        video_df = pd.read_sql(video_query.statement,db.session.bind)
        df = df.merge(video_df, on='Image ID', how='left')
        video_df = None

    # Create annotator 
    if 'Annotator' in required_columns:
        Parent = alias(User)
        annotator_query = db.session.query(
                                        Cluster.id.label('Cluster ID'),
                                        func.coalesce(
                                            Parent.c.username,
                                            User.username,
                                            'AI'
                                        ).label('Annotator') \
                                    )\
                                    .outerjoin(User,User.id==Cluster.user_id)\
                                    .outerjoin(Parent,Parent.c.id==User.parent_id)\
                                    .filter(Cluster.task_id==task_id)

        if trapgroup_id: annotator_query = annotator_query.join(Image,Cluster.images).join(Camera).filter(Camera.trapgroup_id==trapgroup_id)

        annotator_df = pd.read_sql(annotator_query.statement,db.session.bind)
        df = df.merge(annotator_df, on='Cluster ID', how='left')
        annotator_df = None
        df['Annotator'] = df['Annotator'].replace({'Admin': 'AI'}).astype('category')

    # Replace flank with text & capitalize
    if 'Flank' in required_columns:
        flank_query = db.session.query(
                                    Detection.id.label('Sighting ID'),
                                    Detection.flank.label('Flank')
                                )\
                                .join(Image)\
                                .join(Camera)
        
        if trapgroup_id:
            flank_query = flank_query.filter(Camera.trapgroup_id==trapgroup_id)
        else:
            flank_query = flank_query.join(Trapgroup).filter(Trapgroup.survey_id==task.survey_id)

        flank_df = pd.read_sql(flank_query.statement,db.session.bind)
        df = df.merge(flank_df, on='Sighting ID', how='left')
        flank_df = None
        df['Flank'] = df['Flank'].replace(Config.FLANK_TEXT).str.capitalize().astype('category')

    #Remove nulls
    for col in df.select_dtypes(include=['category']).columns: 
        if 'None' not in df[col].cat.categories: df[col] = df[col].cat.add_categories('None')  # Add the 'None' category otherwise fillna will not work

    df.fillna('None', inplace=True)

    #Replace nothings
    df['Species'] = df['Species'].replace({'Nothing': 'None'}, regex=True)

    # drop unneccessary survey column if not needed
    if 'Survey Name' in required_columns:
        df['Survey Name'] = task.survey.name
        df = df.astype({"Survey Name": 'category'})
    
    df.sort_values(by=['Site ID', 'Timestamp'], inplace=True, ascending=True)

    if selectedLevel == 'Unique Capture' or any('Capture' in col for col in required_columns):
        #Add capture ID
        df['Capture'] = df.drop_duplicates(subset=['Camera ID','Timestamp']).groupby('Camera ID').cumcount()+1
        df['Capture'].fillna(0, inplace=True)
        df['Capture'] = df.groupby(['Camera ID','Timestamp'])['Capture'].transform('max')
        df = df.astype({"Capture": 'int32'})

        #Add unique capture ID for labels
        df['Unique Capture ID'] = df.apply(lambda x: str(x['Camera ID']) + '/' + str(x['Capture']), axis=1)

        #Rename Capture column to Capture Number
        df.rename(columns={'Capture':'Capture Number'}, inplace=True)

        if 'Capture ID' in required_columns:
            # Copy Image ID to Capture ID
            df['Capture ID'] = df['Image ID']

    # Count the detections for each label in each required level
    if 'Sighting Count' in required_columns:
        detection_count_levels = [selectedLevel]
        if any(level in detection_count_levels for level in ['Cluster','Unique Capture']) and ('Image' not in detection_count_levels): detection_count_levels.insert(0, 'Image')
        for level in detection_count_levels:
            level_name = level
            level = level + ' ID'
            
            if level_name in ['Cluster','Unique Capture']:
                # Here we want the minimum number of animals in the cluster/capture
                df[level_name+'_sighting_count'] = df.groupby([level, 'Species'])['Image_sighting_count'].transform('max')
            else:
                # here we want the count of the labels in that level
                df[level_name+'_sighting_count'] = df[df['Species']!='None'].groupby([level, 'Species'])['Sighting ID'].transform('nunique')
            
            # propogate all the values to all the rows
            df[level_name+'_sighting_count'] = df.groupby([level, 'Species'])[level_name+'_sighting_count'].transform('max').fillna(0).astype('int32')
        if (selectedLevel!='Image') and ('Image_sighting_count' in df.columns): del df['Image_sighting_count']
        df.rename(columns={level_name+'_sighting_count':'Sighting Count'},inplace=True)

    #Generate necessary urls
    url_levels = []
    for column in required_columns:
        if 'URL' in column:
            url_levels.append(re.split(' URL',column)[0])

    rootUrl = 'https://' + Config.DNS + '/imageViewer?type='
    for level in url_levels:
        level_name = level
        if level == 'Capture':
            level = 'Image ID'
        else:
            level = level + ' ID'
        level_type = level_name.lower()
        if level_type=='site': level_type='trapgroup'
        df[level_name+' URL'] = df.apply(lambda x: generate_url(rootUrl,level_type,x[level]), axis=1)


    return df

# def addChildrenLabels(currentIndex,columns,parentLabels,allLevel,selectedTasks):
#     '''
#     Helper function for generate_csv. Adds the children labels to the dataframe column list, in the correct order.

#         Parameters:
#             currentIndex (int): The index of the column list were the child label columns need to be added
#             columns (list): The column list into which the child label columsn must be added
#             parentLabels (list): The list of labels names that must be added to the column list alongside their children
#             allLevel (str): The level of abstration for which species counts are being added
#             selectedTasks (list): The task IDs for which the csv has been requested

#         Returns:
#             currentIndex (int): The last index where columns were inserted
#             columns (list): The updated version of the column list
#     '''
    
#     Parent = alias(Label)

#     for parentLabel in parentLabels:

#         column = allLevel+'_'+parentLabel.replace(' ','_').lower()+'_count'
#         if column not in columns:
#             columns.insert(currentIndex,column)
#         currentIndex += 1

#         childrenLabels =    [ r[0] for r in 
#                                 db.session.query(Label.description)\
#                                 .join(Parent,Parent.c.id==Label.parent_id)\
#                                 .filter(Label.task_id.in_(selectedTasks))\
#                                 .filter(Parent.c.description==parentLabel)\
#                                 .all()
#                             ]

#         if childrenLabels:
#             currentIndex, columns = addChildrenLabels(currentIndex,columns,childrenLabels,allLevel,selectedTasks)

#     return currentIndex, columns

def handle_custom_columns(columns,row,custom_split):
    '''
    Helper function for generate_csv. Returns the custom column value for a row of data.

        Parameters:
            columns (list): The column names of the row of data
            row (dictionary): The row of data itself
            custom_split (list): The custom column information list

        Returns:
            output (string): The custom column value for the input row.
    '''
	
    output = ''
    for item in custom_split:
        if item in columns:
            output += str(row[item])
        else:
            output += str(item)
    return output

def combine_list(list):
    reply='['
    for item in list:
        reply += str(item)+';'
    reply = reply[:-1]
    reply += ']'
    return reply

@celery.task(bind=True,max_retries=5,ignore_result=True)
def generate_csv_parent(self,selectedTasks, selectedLevel, requestedColumns, custom_columns, label_type, includes, excludes, startDate, endDate, column_translations, user_name, download_id):
    '''
    Celery task for generating a csv file. Locally saves a csv file for the requested tasks, with the requested column and row information.

        Paramters:
            selectedTasks (list): List of task IDs for which a csv is required
            selectedLevel (string): The level of abstration each row in the csv is to represent (image, cluster, camera etc.)
            requestedColumns (list): List of columns for the csv, in the order in which they are required
            custom_columns (list): List of the descriptions of custom columns requested in the csv
            label_type (str): The type of labelling scheme to be uses - row, column or list
            includes (list): List of label names that should included
            excludes (list): List of label names that should excluded
            startDate (dateTime): The start date for the data to be included in the csv
            endDate (dateTime): The end date for the data to be included in the csv
            user_name (str): The name of the user that has requested the csv
            download_id (int): The id of the download request in the database
    '''
    try: 
        task = db.session.query(Task).get(selectedTasks[0])
        filePath = task.survey.organisation.folder+'/docs/'
        randomness = randomString()
        fileName = task.survey.organisation.name+'_'+user_name+'_'+task.survey.name+'_'+task.name+'_'+randomness+'.csv'

        if selectedLevel == 'Capture': selectedLevel = 'Unique Capture'

        # Generate include an exclude lists
        include = [r[0] for r in db.session.query(Label.id).filter(Label.task_id.in_(selectedTasks)).filter(Label.description.in_(includes)).distinct().all()]
        include.extend([r[0] for r in db.session.query(Label.id).filter(Label.task_id==None).filter(Label.description.in_(includes)).distinct().all()])
        exclude = [r[0] for r in db.session.query(Label.id).filter(Label.task_id.in_(selectedTasks)).filter(Label.description.in_(excludes)).distinct().all()]
        exclude.extend([r[0] for r in db.session.query(Label.id).filter(Label.task_id==None).filter(Label.description.in_(excludes)).distinct().all()])

        # Handle bounding boxes
        if 'Boxes' in requestedColumns:
            index = requestedColumns.index('Boxes')
            requestedColumns[index:index] = ['Left','Right','Top','Bottom','Score']
            requestedColumns.remove('Boxes')

        required_columns = requestedColumns.copy()
        for c in custom_columns[str(selectedTasks[0])]:
            custom = custom_columns[str(selectedTasks[0])][c]
            custom_split = [r for r in re.split('%%%%',custom) if r != '']
            required_columns = required_columns + [c for c in custom_split if c not in required_columns]

        overall_det_count = rDets(db.session.query(Detection).join(Image).join(Camera).join(Trapgroup).filter(Trapgroup.survey_id==task.survey_id))
        if include: overall_det_count = overall_det_count.join(Labelgroup).join(Label,Labelgroup.labels).filter(Labelgroup.task_id.in_(selectedTasks)).filter(Label.id.in_(include))
        if exclude: overall_det_count = overall_det_count.join(Labelgroup).join(Label,Labelgroup.labels).filter(Labelgroup.task_id.in_(selectedTasks)).filter(~Label.id.in_(exclude))
        if startDate: overall_det_count = overall_det_count.filter(Image.corrected_timestamp>=startDate)
        if endDate: overall_det_count = overall_det_count.filter(Image.corrected_timestamp<=endDate)
        overall_det_count = overall_det_count.distinct().count()

        if overall_det_count<200000:
            # Small csv, it will be faster to just generate it locally
            generate_csv(
                trapgroup_id=None,
                task_id=selectedTasks[0],
                filename=filePath+fileName,
                selectedLevel=selectedLevel,
                requestedColumns=requestedColumns,
                custom_columns=custom_columns,
                label_type=label_type,
                include=include,
                exclude=exclude,
                startDate=startDate,
                endDate=endDate,
                column_translations=column_translations,
                required_columns=required_columns
            )

        else:
            # Large csv, let's parallelise
            trapgroup_data = db.session.query(Trapgroup.id,func.count(Detection.id)).filter(Trapgroup.survey_id==task.survey_id).join(Camera).join(Image).join(Detection).group_by(Trapgroup.id).all()
            trapgroup_filenames = []
            results = []
            for trapgroup_id, count in trapgroup_data:
                trapgroup_filename = task.survey.organisation.name+'_'+user_name+'_'+task.survey.name+'_'+str(trapgroup_id)+'_'+randomness+'.csv'
                trapgroup_filenames.append(trapgroup_filename)
                kwargs = {
                    'trapgroup_id':trapgroup_id,
                    'task_id':selectedTasks[0],
                    'filename':trapgroup_filename,
                    'selectedLevel':selectedLevel,
                    'requestedColumns':requestedColumns,
                    'custom_columns':custom_columns,
                    'label_type':label_type,
                    'include':include,
                    'exclude':exclude,
                    'startDate':startDate,
                    'endDate':endDate,
                    'column_translations':column_translations,
                    'required_columns':required_columns
                }
                if count < 400000:
                    results.append(generate_csv.apply_async(kwargs=kwargs, queue='parallel'))
                else:
                    results.append(generate_csv.apply_async(kwargs=kwargs, queue='ram_intensive'))
            
            if results:
                #Wait for processing to complete
                db.session.remove()
                GLOBALS.lock.acquire()
                with allow_join_result():
                    for result in results:
                        try:
                            result.get()
                        except Exception:
                            app.logger.info(' ')
                            app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
                            app.logger.info(traceback.format_exc())
                            app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
                            app.logger.info(' ')
                        
                        result.forget()
                GLOBALS.lock.release()

            with tempfile.NamedTemporaryFile(delete=True, suffix='.csv') as temp_output_file:
                with open(temp_output_file.name, "w", newline='') as outfile:
                    writer = None

                    for trapgroup_filename in trapgroup_filenames:
                        with tempfile.NamedTemporaryFile(delete=True, suffix='.csv') as temp_in_file:
                            GLOBALS.s3client.download_file(Bucket=Config.BUCKET, Key=task.survey.organisation.folder+'-comp/csvs/'+trapgroup_filename, Filename=temp_in_file.name)
                            
                            with open(temp_in_file.name, "r", newline='') as infile:
                                reader = csv.reader(infile)

                                # Get header from first file
                                if writer is None:
                                    writer = csv.writer(outfile)
                                    writer.writerow(next(reader))  # Write header

                                else:
                                    next(reader)  # Skip header for subsequent files

                                # Stream content
                                writer.writerows(reader)

                # Save to S3
                GLOBALS.s3client.upload_file(Filename=temp_output_file.name, Bucket=Config.BUCKET, Key=filePath+fileName)

            # Delete trapgroup files
            for trapgroup_filename in trapgroup_filenames:
                key = task.survey.organisation.folder+'-comp/csvs/'+trapgroup_filename
                GLOBALS.s3client.delete_object(Bucket=Config.BUCKET, Key=key)

        # Set request status to complete
        download_request = db.session.query(DownloadRequest).get(download_id)
        if download_request:
            download_request.status = 'Available'
            download_request.timestamp = datetime.now() + timedelta(days=7)
            download_request.name = randomness
            db.session.commit()

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        self.retry(exc=exc, countdown= retryTime(self.request.retries))

    finally:
        db.session.remove()

    return True

@celery.task(bind=True,max_retries=5)
def generate_csv(self,trapgroup_id,task_id,filename,selectedLevel,requestedColumns,custom_columns,label_type,include,exclude,startDate,endDate,column_translations,required_columns):
    ''' Generate a dataframe for a trapgroup and save to CSV.'''
    
    try:
        task = db.session.query(Task).get(task_id)

        if trapgroup_id:
            path = task.survey.organisation.folder+'-comp/csvs/'+filename
        else:
            path = filename

        df = create_task_dataframe(task_id,selectedLevel,include,exclude,trapgroup_id,startDate,endDate,required_columns)

        if (df is not None) and (len(df) > 0):
            for custom_name in custom_columns[str(task_id)]:
                custom = custom_columns[str(task_id)][custom_name]
                custom_split = [r for r in re.split('%%%%',custom) if r != '']
                df[custom_name] = df.apply(lambda x: handle_custom_columns(df.columns,x,custom_split), axis=1)

            if selectedLevel=='Sighting': df = df[df['Sighting ID']!='None']

            if label_type=='list':

                # Aggregate labels and their sighting counts by splitting them out to their own df
                agg_rules = {'Species': lambda x: '|'.join(x)}
                if 'Sighting Count' in df.columns: agg_rules['Sighting Count'] = lambda x: '|'.join(map(str, x))
                cols = [selectedLevel+' ID','Species']
                if 'Sighting Count' in df.columns: cols.append('Sighting Count')
                df_grouped = df[df['Species']!='None'].groupby([selectedLevel+' ID', 'Species']).first().reset_index()[cols]
                del df['Species']
                if 'Sighting Count' in df.columns: del df['Sighting Count']
                df_grouped = df_grouped.groupby(selectedLevel+' ID').agg(agg_rules).reset_index()

                # Aggregate Tags, Individuals & rest of the columns
                # Exclude None value if there are other tags or individuals
                agg_rules = {'Informational Tags': lambda x: ('|'.join(x[x!='None'].unique()) if len(x[x!='None']) > 0 else 'None')}
                if 'Individuals' in df.columns: agg_rules['Individuals'] = lambda x: ('|'.join(x[x!='None'].unique()) if len(x[x!='None']) > 0 else 'None')
                for column in [column for column in df.columns if (column not in agg_rules.keys()) and (column != selectedLevel+' ID')]: agg_rules[column] = 'first'
                df = df.groupby(selectedLevel+' ID', sort=False).agg(agg_rules).reset_index()

                # Merge the two back together again
                df = df.merge(df_grouped, on=selectedLevel+' ID', how='left')
                df['Species'].fillna('None', inplace=True)
                if 'Sighting Count' in df.columns: df['Sighting Count'].fillna(0, inplace=True)

            else:
                subset_list = [selectedLevel+' ID','Species']
                if 'Informational Tags' in requestedColumns: subset_list.append('Informational Tags')
                if 'Individuals' in requestedColumns: subset_list.append('Individuals')
                df = df.drop_duplicates(subset=subset_list, keep='first')
                # We only want to include the None Species rows where there is no other species in the selectedLevel
                only_none = df.groupby(selectedLevel+' ID')['Species'].transform(lambda x: (x == 'None').all())
                df = df[only_none | (df['Species'] != 'None')]

            df = df[requestedColumns]
        
        else:
            df = pd.DataFrame(columns=requestedColumns)
        
        # column translation:
        levels = ['image', 'capture', 'cluster', 'camera', 'site', 'trapgroup', 'survey']
        changes = {}
        for column in df.columns:
            for translation in column_translations:
                if translation.lower() == column.lower():
                    changes[column] = column.replace(translation,column_translations[translation])
                elif translation in column and column[-1].isdigit() and translation.lower() not in levels:
                    changes[column] = column.replace(translation,column_translations[translation])

        if len(changes) != 0:
            df.rename(columns=changes,inplace=True)
        
        with tempfile.NamedTemporaryFile(delete=True, suffix='.csv') as temp_file:
            df.to_csv(temp_file.name, index=False)
            GLOBALS.s3client.upload_file(Filename=temp_file.name, Bucket=Config.BUCKET, Key=path)

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        self.retry(exc=exc, countdown= retryTime(self.request.retries))

    finally:
        db.session.remove()

    return True

@celery.task(bind=True,max_retries=5,ignore_result=True)
def generate_wildbook_export(self,task_id, data, user_name, download_request_id):
    '''
    Celery task for generating the WildBook export format. Saves export in zip file locally.

        Parameters:
            task_id (int): The task for which the export is required
            data (dict): The required data for the export - species, genus, epithet, wildbookid
    '''
    
    try:
        os.makedirs('docs', exist_ok=True)
        task = db.session.query(Task).get(task_id)
        # fileName = task.survey.user.folder+'/docs/'+task.survey.user.username+'_'+task.survey.name+'_'+task.name
        randomness = randomString()
        fileName = task.survey.organisation.folder+'/docs/'+task.survey.organisation.name+'_'+user_name+'_'+task.survey.name+'_'+task.name + '_' + randomness
        # # Delete old file if exists
        # try:
        #     GLOBALS.s3client.delete_object(Bucket=Config.BUCKET, Key=fileName+'.zip')
        # except:
        #     pass

        tempFolderName = fileName+'_temp'
        species = db.session.query(Label).get(int(data['species']))

        df = pd.read_sql(db.session.query( \
                        Detection.id.label('detection'), \
                        Camera.path.label('path'), \
                        Camera.id.label('camera_id'), \
                        Image.filename.label('filename'), \
                        Image.corrected_timestamp.label('timestamp'), \
                        Trapgroup.latitude.label('Encounter.decimalLatitude'), \
                        Trapgroup.longitude.label('Encounter.decimalLongitude')) \
                        .join(Image, Image.id==Detection.image_id) \
                        .join(Camera, Camera.id==Image.camera_id) \
                        .join(Trapgroup, Trapgroup.id==Camera.trapgroup_id) \
                        .join(Labelgroup, Labelgroup.detection_id==Detection.id) \
                        .filter(Labelgroup.task_id==task_id) \
                        .filter(Labelgroup.labels.contains(species)) \
                        .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS)) \
                        .filter(Detection.static == False) \
                        .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES)) \
                        .statement,db.session.bind)

        if df.empty:
            with zipfile.ZipFile(fileName+'.zip', 'w') as zipf:
                pass
            GLOBALS.s3client.upload_file(fileName+'.zip', Config.BUCKET, fileName+'.zip')
            os.remove(fileName+'.zip')
        else:
            df['Encounter.genus'] = data['genus'][:1].upper() + data['genus'][1:].lower()
            df['Encounter.specificEpithet'] = data['epithet'].lower()
            df['Encounter.submitterID'] = data['wildbookid']

            df['Encounter.year'] = df.apply(lambda x: x.timestamp.year if x.timestamp else None, axis=1)
            df['Encounter.month'] = df.apply(lambda x: x.timestamp.month if x.timestamp else None, axis=1)
            df['Encounter.day'] = df.apply(lambda x: x.timestamp.day if x.timestamp else None, axis=1)
            df['Encounter.hour'] = df.apply(lambda x: x.timestamp.hour if x.timestamp else None, axis=1)
            df['Encounter.minutes'] = df.apply(lambda x: x.timestamp.minute if x.timestamp else None, axis=1)

            df['Encounter.mediaAsset0'] = df.apply(lambda x: (str(x.camera_id)+'_'+x.filename).replace('(','_').replace(')','_'), axis=1)
            df = df.drop_duplicates(subset=['Encounter.mediaAsset0'], keep='first').reset_index()

            del df['timestamp']
            del df['detection']
            del df['camera_id']
            del df['index']

            df = df.fillna('None')

            df['comp_path'] = df.apply(lambda x: x['path'].split('/',1)[0] + '-comp/' + x['path'].split('/',1)[1], axis=1)

            if os.path.isdir(tempFolderName):
                shutil.rmtree(tempFolderName)
            os.makedirs(tempFolderName, exist_ok=True)

            for n in range((math.ceil(len(df)/1000))):
                subsetFolder = tempFolderName+'/'+str(n+1)
                os.makedirs(subsetFolder, exist_ok=True)
                tempDF = df.loc[n*1000:(n+1)*1000-1]

                tempDF.apply(lambda x: file_download_for_export(x,subsetFolder), axis=1)

                del tempDF['path']
                del tempDF['filename']
                del tempDF['comp_path']
                tempDF.to_csv(subsetFolder+'/metadata.csv',index=False)

            shutil.make_archive(tempFolderName, 'zip', tempFolderName)
            GLOBALS.s3client.upload_file(tempFolderName+'.zip', Config.BUCKET, fileName+'.zip')
            shutil.rmtree(tempFolderName)

        # Schedule deletion
        # deleteFile.apply_async(kwargs={'fileName': fileName+'.zip'}, countdown=3600)

        download_request = db.session.query(DownloadRequest).get(download_request_id)
        if download_request:
            download_request.status = 'Available'
            download_request.name = randomness
            db.session.commit()

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        self.retry(exc=exc, countdown= retryTime(self.request.retries))

    finally:
        db.session.remove()

    return True

def file_download_for_export(x,subsetFolder):
    ''' Tries to download the file for a row of data. Uses comp file if raw not available'''
    try:
        GLOBALS.s3client.download_file(
            Bucket=Config.BUCKET, 
            Key=(x['path']+'/'+x['filename']), 
            Filename=(subsetFolder+'/'+x['Encounter.mediaAsset0'])
        )
    except:
        try:
            GLOBALS.s3client.download_file(
                Bucket=Config.BUCKET, 
                Key=(x['comp_path']+'/'+x['filename']), 
                Filename=(subsetFolder+'/'+x['Encounter.mediaAsset0'])
            )
        except:
            pass

def num_to_excel_col(n):
    '''Converts a numerical column index into an Excel format column.'''
    if n < 1:
        raise ValueError("Number must be positive")
    result = ""
    while True:
        if n > 26:
            n, r = divmod(n - 1, 26)
            result = chr(r + ord('A')) + result
        else:
            return chr(n + ord('A') - 1) + result


def child_headings(task_id, label, currentCol, speciesColumns, sheet):
    '''
    Helper function for generate_excel. Generates the species columns for the excel file in the correct order.

        Parameters:
            task_id (int): The task for which the Excel file is being prepared
            label (Label): The label for which child labels must be added
            currentCol (int): Column index
            speciesColumns (list): The species columns list
            sheet (openpyxl.sheet): The Excel sheet object beiong written to

        Returns:
            currentCol (int): The updated index
            speciesColumns (list): The updated species column list
            sheet (openpyxl.sheet): The updated Excel sheet objec
    '''
    
    children = db.session.query(Label).filter(Label.task_id==task_id).filter(Label.parent==label).all()
    for child in children:
        currentColumn = num_to_excel_col(currentCol)
        speciesColumns[child.description] = currentColumn
        currentColumn += '3'
        sheet[currentColumn] = child.description
        sheet[currentColumn].border = Border(bottom=Side(style='thin'))
        currentCol += 1

        childChildren = db.session.query(Label).filter(Label.task_id==task_id).filter(Label.parent==child).all()
        if len(childChildren) > 0:
            sheet[currentColumn[:-1]+'2'] = label.description
            sheet[currentColumn] = 'Unknown'
            sheet[currentColumn].border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
            currentCol, speciesColumns, sheet = child_headings(task_id, child, currentCol, speciesColumns, sheet)
    return currentCol, speciesColumns, sheet


@celery.task(bind=True,max_retries=5,ignore_result=True)
def generate_excel(self,task_id,user_name,download_id):
    '''Celery task for generating an Excel summary of a specified task. Saves the file locally for download.'''
    try:
        app.logger.info('generate_excel commenced')
        task = db.session.query(Task).get(task_id)
        survey = task.survey
        # user = survey.user
        organisation = survey.organisation

        right_border = Border(right=Side(style='thin'))
        left_border = Border(left=Side(style='thin'))
        top_border = Border(top=Side(style='thin'))
        bottom_border = Border(bottom=Side(style='thin'))

        greyFill = PatternFill(start_color='DCDCDC',
                        end_color='DCDCDC',
                        fill_type='solid')

        whiteFill = PatternFill(start_color='FFFFFF',
                        end_color='FFFFFF',
                        fill_type='solid')

        workbook = Workbook()
        sheet = workbook.active

        #Headings
        sheet["A1"] = survey.name
        sheet["A1"].font = Font(size = "18", bold=True)
        sheet.row_dimensions[1].height = 25
        # sheet.merge_cells('A1:E1')
        sheet["B2"] = 'Co-ordinates'
        sheet["A3"] = 'Site'
        sheet["A3"].border = bottom_border
        sheet["B3"] = 'Longitude'
        sheet["B3"].border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
        sheet["C3"] = 'Latitude'
        sheet["C3"].border = bottom_border
        sheet["D3"] = 'Altitude'
        sheet["D3"].border = Border(right=Side(style='thin'), bottom=Side(style='thin'))
        sheet["E3"] = 'Total Clusters'
        sheet["E3"].border = Border(right=Side(style='thin'), bottom=Side(style='thin'))


        speciesColumns = {}
        currentCol = 6
        parentLabels = db.session.query(Label).filter(Label.task_id==task_id).filter(~Label.parent.has()).all()
        vhl = db.session.query(Label).get(GLOBALS.vhl_id)
        parentLabels.append(vhl)
        for label in parentLabels:
            currentColumn = num_to_excel_col(currentCol)
            speciesColumns[label.description] = currentColumn
            currentColumn += '3'
            sheet[currentColumn] = label.description
            sheet[currentColumn].border = bottom_border
            currentCol += 1

            children = db.session.query(Label).filter(Label.task_id==task_id).filter(Label.parent==label).all()
            if len(children)>0:
                sheet[currentColumn[:-1]+'2'] = label.description
                sheet[currentColumn] = 'Unknown'
                sheet[currentColumn].border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
                currentCol, speciesColumns, sheet = child_headings(task_id, label, currentCol, speciesColumns, sheet)

        finalColumn = currentColumn[:-1]
        sheet[finalColumn + str(3)].border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
        if Config.DEBUGGING: print(speciesColumns)

        #Generate Rows
        traps = db.session.query(Trapgroup).filter(Trapgroup.survey==survey).all()
        labels = db.session.query(Label).filter(Label.task_id==task_id).all()
        labels.append(vhl)
        trapgroupsCompleted = []
        currentRow = 4
        currentFill = whiteFill
        for trapgroup in traps:
            if trapgroup.tag not in trapgroupsCompleted:
                sheet["A"+str(currentRow)] = trapgroup.tag
                sheet["B"+str(currentRow)] = trapgroup.longitude
                sheet["C"+str(currentRow)] = trapgroup.latitude
                sheet["D"+str(currentRow)] = trapgroup.altitude
                trapgroupsCompleted.append(trapgroup.tag)
                if currentFill == greyFill:
                    currentFill = whiteFill
                else:
                    currentFill = greyFill
            sheet["A"+str(currentRow)].fill = currentFill
            sheet["B"+str(currentRow)].fill = currentFill
            sheet["B"+str(currentRow)].border = Border(left=Side(style='thin'))
            sheet["C"+str(currentRow)].fill = currentFill
            sheet["D"+str(currentRow)].fill = currentFill
            sheet["D"+str(currentRow)].border = Border(right=Side(style='thin'))

            camera_ids = [camera.id for camera in trapgroup.cameras]

            sheet["E"+str(currentRow)] = db.session.query(Cluster) \
                                                .join(Image, Cluster.images) \
                                                .filter(Cluster.task_id==task_id) \
                                                .filter(Image.camera_id.in_(camera_ids)) \
                                                .distinct(Cluster.id) \
                                                .count()
            sheet["E"+str(currentRow)].fill = currentFill
            sheet["E"+str(currentRow)].border = Border(right=Side(style='thin'))
            
            for label in labels:
                clusterCount = db.session.query(Cluster) \
                                    .join(Image, Cluster.images) \
                                    .join(Detection)\
                                    .join(Labelgroup)\
                                    .filter(Labelgroup.task_id==task_id)\
                                    .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                                    .filter(Detection.static==False)\
                                    .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                                    .filter(Cluster.task_id==task_id) \
                                    .filter(Labelgroup.labels.contains(label)) \
                                    .filter(Image.camera_id.in_(camera_ids)) \
                                    .distinct(Cluster.id) \
                                    .count()

                if clusterCount > 0:
                    sheet[speciesColumns[label.description]+str(currentRow)] = clusterCount

                labelChildren = db.session.query(Label).filter(Label.parent==label).filter(Label.task==task).first()
                if labelChildren:
                    sheet[speciesColumns[label.description]+str(currentRow)].border = left_border

                sheet[speciesColumns[label.description]+str(currentRow)].fill = currentFill

            sheet[finalColumn+str(currentRow)].border = Border(right=Side(style='thin'), left=Side(style='thin'))
            currentRow += 1

        sheet['A'+str(currentRow)].border = Border(top=Side(style='thin'))
        sheet['B'+str(currentRow)].border = Border(top=Side(style='thin'))
        sheet['C'+str(currentRow)].border = Border(top=Side(style='thin'))
        sheet['D'+str(currentRow)].border = Border(top=Side(style='thin'))
        sheet['E'+str(currentRow)].border = Border(top=Side(style='thin'))

        for label in labels:
            sheet[speciesColumns[label.description]+str(currentRow)].border = Border(top=Side(style='thin'))
            sheet[speciesColumns[label.description]+str(currentRow)].fill = whiteFill
            sheet[speciesColumns[label.description]+str(currentRow+3)].fill = whiteFill
            sheet[speciesColumns[label.description]+str(currentRow+6)].fill = whiteFill
            sheet[speciesColumns[label.description]+str(currentRow+7)].fill = whiteFill
            sheet[speciesColumns[label.description]+'1'].fill = whiteFill
            sheet[speciesColumns[label.description]+'2'].fill = whiteFill
            sheet[speciesColumns[label.description]+'3'].fill = whiteFill

        for item in ['A','B','C','D','E']:
            sheet[item+str(currentRow)].fill = whiteFill
            sheet[item+str(currentRow+3)].fill = whiteFill
            sheet[item+str(currentRow+6)].fill = whiteFill
            sheet[item+str(currentRow+7)].fill = whiteFill
            sheet[item+'1'].fill = whiteFill
            sheet[item+'2'].fill = whiteFill
            sheet[item+'3'].fill = whiteFill

        sheet[finalColumn+str(currentRow-1)].border = Border(right=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))

        #Generate Totals
        currentRow += 1
        sheet["A"+str(currentRow)] = 'Totals'
        sheet["A"+str(currentRow)].border = Border(right=Side(style='thin'), left=Side(style='thin'), top=Side(style='thin'))
        sheet["A"+str(currentRow)].fill = greyFill
        sheet.merge_cells('A'+str(currentRow)+':D'+str(currentRow))
        sheet["A"+str(currentRow+1)] = 'Percentage'
        sheet["A"+str(currentRow+1)].border = Border(right=Side(style='thin'), left=Side(style='thin'), bottom=Side(style='thin'))
        sheet["A"+str(currentRow+1)].fill = whiteFill
        sheet.merge_cells('A'+str(currentRow+1)+':D'+str(currentRow+1))
        totalCount = db.session.query(Cluster).filter(Cluster.task_id==task_id).distinct(Cluster.id).count()
        sheet["E"+str(currentRow)] = totalCount
        sheet["E"+str(currentRow)].border = top_border + right_border
        sheet["E"+str(currentRow)].fill = greyFill
        sheet["E"+str(currentRow+1)].border = bottom_border + right_border
        sheet["E"+str(currentRow+1)].fill = whiteFill
        for label in labels:
            clusterCount = db.session.query(Cluster) \
                                    .join(Image, Cluster.images) \
                                    .join(Detection)\
                                    .join(Labelgroup)\
                                    .filter(Labelgroup.task_id==task_id)\
                                    .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                                    .filter(Detection.static==False)\
                                    .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                                    .filter(Cluster.task_id==task_id) \
                                    .filter(Labelgroup.labels.contains(label)) \
                                    .distinct(Cluster.id) \
                                    .count()

            sheet[speciesColumns[label.description]+str(currentRow)] = clusterCount
            sheet[speciesColumns[label.description]+str(currentRow)].border = top_border
            sheet[speciesColumns[label.description]+str(currentRow)].fill = greyFill
            if totalCount > 0:
                sheet[speciesColumns[label.description]+str(currentRow+1)] = round((clusterCount/totalCount)*100,2)
            else:
                sheet[speciesColumns[label.description]+str(currentRow+1)] = 0
            sheet[speciesColumns[label.description]+str(currentRow+1)].border = bottom_border
            sheet[speciesColumns[label.description]+str(currentRow+1)].fill = whiteFill

            labelChildren = db.session.query(Label).filter(Label.parent==label).filter(Label.task==task).first()
            if labelChildren:
                sheet[speciesColumns[label.description]+str(currentRow)].border = Border(left=Side(style='thin'), top=Side(style='thin'))
                sheet[speciesColumns[label.description]+str(currentRow+1)].border = Border(left=Side(style='thin'), bottom=Side(style='thin'))

        sheet[finalColumn+str(currentRow)].border = Border(right=Side(style='thin'), top=Side(style='thin'), left=Side(style='thin'))
        sheet[finalColumn+str(currentRow+1)].border = Border(right=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))

        #Generate Category Totals
        currentRow += 3
        sheet["A"+str(currentRow)] = 'Category Totals'
        sheet["A"+str(currentRow)].border = Border(right=Side(style='thin'), left=Side(style='thin'), top=Side(style='thin'))
        sheet["A"+str(currentRow)].fill = greyFill
        sheet.merge_cells('A'+str(currentRow)+':D'+str(currentRow))
        sheet["A"+str(currentRow+1)] = 'Category Percentages'
        sheet["A"+str(currentRow+1)].border = Border(right=Side(style='thin'), left=Side(style='thin'), bottom=Side(style='thin'))
        sheet["A"+str(currentRow+1)].fill = whiteFill
        sheet.merge_cells('A'+str(currentRow+1)+':D'+str(currentRow+1))
        parentLabels.append(vhl)
        sheet["E"+str(currentRow)] = totalCount
        sheet["E"+str(currentRow)].border = top_border + right_border
        sheet["E"+str(currentRow)].fill = greyFill
        sheet["E"+str(currentRow+1)].border = bottom_border + right_border
        sheet["E"+str(currentRow+1)].fill = whiteFill
        for label in parentLabels:
            count = 0
            count += db.session.query(Cluster) \
                            .join(Image, Cluster.images) \
                            .join(Detection)\
                            .join(Labelgroup)\
                            .filter(Labelgroup.task_id==task_id)\
                            .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                            .filter(Detection.static==False)\
                            .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                            .filter(Cluster.task_id==task_id) \
                            .filter(Labelgroup.labels.contains(label)) \
                            .distinct(Cluster.id) \
                            .count()
            
            labelChildren = db.session.query(Label).filter(Label.parent==label).filter(Label.task==task).all()
            for child in labelChildren:
                count += db.session.query(Cluster) \
                            .join(Image, Cluster.images) \
                            .join(Detection)\
                            .join(Labelgroup)\
                            .filter(Labelgroup.task_id==task_id)\
                            .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                            .filter(Detection.static==False)\
                            .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                            .filter(Cluster.task_id==task_id) \
                            .filter(Labelgroup.labels.contains(child)) \
                            .distinct(Cluster.id) \
                            .count()

            sheet[speciesColumns[label.description]+str(currentRow)] = count
            if totalCount > 0:
                sheet[speciesColumns[label.description]+str(currentRow+1)] = round((count/totalCount)*100,2)
            else:
                sheet[speciesColumns[label.description]+str(currentRow+1)] = 0

        for label in labels:
            sheet[speciesColumns[label.description]+str(currentRow)].border = top_border
            sheet[speciesColumns[label.description]+str(currentRow)].fill = greyFill
            sheet[speciesColumns[label.description]+str(currentRow+1)].border = bottom_border
            sheet[speciesColumns[label.description]+str(currentRow+1)].fill = whiteFill

            labelChildren = db.session.query(Label).filter(Label.parent==label).filter(Label.task==task).first()
            if labelChildren:
                sheet[speciesColumns[label.description]+str(currentRow)].border = Border(left=Side(style='thin'), top=Side(style='thin'))
                sheet[speciesColumns[label.description]+str(currentRow+1)].border = Border(left=Side(style='thin'), bottom=Side(style='thin'))

        sheet[finalColumn+str(currentRow)].border = Border(right=Side(style='thin'), top=Side(style='thin'), left=Side(style='thin'))
        sheet[finalColumn+str(currentRow+1)].border = Border(right=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))

        # fileName = user.folder+'/docs/'+user.username+'_'+survey.name+'_'+task.name+'.xlsx'
        randomness = randomString()
        fileName = organisation.folder+'/docs/'+organisation.name+'_'+user_name+'_'+survey.name+'_'+task.name+'_'+randomness+'.xlsx'

        # Write new file to S3 for fetching
        with tempfile.NamedTemporaryFile(delete=True, suffix='.xlsx') as temp_file:
            workbook.save(filename=temp_file.name)
            GLOBALS.s3client.put_object(Bucket=Config.BUCKET,Key=fileName,Body=temp_file)

        # Schedule deletion
        # deleteFile.apply_async(kwargs={'fileName': fileName}, countdown=3600)

        download_request = db.session.query(DownloadRequest).get(download_id)
        if download_request:
            download_request.status = 'Available'
            download_request.timestamp = datetime.now() + timedelta(days=7)
            download_request.name = randomness
            db.session.commit()

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        self.retry(exc=exc, countdown= retryTime(self.request.retries))

    finally:
        db.session.remove()

    return None

def get_image_paths_and_labels(image,task,individual_sorted,species_sorted,flat_structure,requested_labels,include_empties):
    '''Returns the paths, labels and tags for a particular image and task for image sorting and EXIF labelling.'''

    if '0' in requested_labels:
        requested_labels = [r.id for r in task.labels]
        requested_labels.append(GLOBALS.vhl_id)
        requested_labels.append(GLOBALS.knocked_id)
        requested_labels.append(GLOBALS.unknown_id)
    if include_empties: requested_labels.append(GLOBALS.nothing_id)
    requested_labels = [int(r) for r in requested_labels]

    splitPath = re.split('/',image.camera.path)

    imageLabels = db.session.query(Label)\
                        .join(Labelgroup,Label.labelgroups)\
                        .join(Detection)\
                        .filter(Detection.image==image)\
                        .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                        .filter(Detection.static==False)\
                        .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                        .filter(Labelgroup.task==task)\
                        .distinct().order_by(Label.description).all()
    
    if imageLabels == []:
        imageLabels = [db.session.query(Label).get(GLOBALS.nothing_id)]

    imageTags = db.session.query(Tag)\
                        .join(Labelgroup,Tag.labelgroups)\
                        .join(Detection)\
                        .filter(Detection.image==image)\
                        .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                        .filter(Detection.static==False)\
                        .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                        .filter(Labelgroup.task==task)\
                        .distinct().order_by(Tag.description).all()

    imagePaths = []
    baseName = image.camera.trapgroup.tag + '_' + stringify_timestamp(image.corrected_timestamp, True)	
    for label in imageLabels:
        
        if (label.id==GLOBALS.nothing_id) and not include_empties:
            continue
        
        individuals = [None]
        
        if individual_sorted:
            individuals = db.session.query(Individual)\
                                .join(Detection,Individual.detections)\
                                .filter(Detection.image==image)\
                                .filter(Individual.tasks.contains(task))\
                                .filter(Individual.species==label.description)\
                                .filter(Individual.active==True)\
                                .distinct().all()
            if len(individuals)==0: individuals = [None]
        
        for individual in individuals:
            if label.id in requested_labels:
                imagePath = ''
                if species_sorted:  imagePath += '/' + label.description.replace('/','_').replace('\\','_')
                if individual and individual_sorted: imagePath += '/' + individual.name
                if flat_structure:
                    filename = baseName
                    for imageLabel in imageLabels: filename += '_' + imageLabel.description.replace(' ','_').replace('/','_').replace('\\','_')
                    filename += '_' + str(image.id) + '.jpg'
                    imagePath += '/' + filename
                else:
                    startPoint = 1
                    if splitPath[1]==task.survey.name: startPoint=2
                    for split in splitPath[startPoint:]: imagePath += '/' + split
                    imagePath += '/' +image.filename
                imagePaths.append(imagePath)

    return list(set(imagePaths)), [label.description for label in imageLabels], [tag.description for tag in imageTags]

def get_video_paths_and_labels(video,task,individual_sorted,species_sorted,flat_structure,requested_labels,include_empties):
    '''Returns the paths, labels and tags for a particular video and task for video sorting and metadata labelling.'''

    if '0' in requested_labels:
        requested_labels = [r.id for r in task.labels]
        requested_labels.append(GLOBALS.vhl_id)
        requested_labels.append(GLOBALS.knocked_id)
        requested_labels.append(GLOBALS.unknown_id)
    if include_empties: requested_labels.append(GLOBALS.nothing_id)
    requested_labels = [int(r) for r in requested_labels]

    splitPath = re.split('/', video.camera.path.split('_video_images_')[0])

    videoImages = db.session.query(Image).filter(Image.camera_id == video.camera_id).all()
    video_images_ids = [image.id for image in videoImages]

    videoLabels = db.session.query(Label)\
                        .join(Labelgroup,Label.labelgroups)\
                        .join(Detection)\
                        .join(Image)\
                        .filter(Image.id.in_(video_images_ids))\
                        .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                        .filter(Detection.static==False)\
                        .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                        .filter(Labelgroup.task==task)\
                        .distinct().order_by(Label.description).all()
    
    if videoLabels == []:
        videoLabels = [db.session.query(Label).get(GLOBALS.nothing_id)]

    videoTags = db.session.query(Tag)\
                        .join(Labelgroup,Tag.labelgroups)\
                        .join(Detection)\
                        .join(Image)\
                        .filter(Image.id.in_(video_images_ids))\
                        .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                        .filter(Detection.static==False)\
                        .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                        .filter(Labelgroup.task==task)\
                        .distinct().order_by(Tag.description).all()
    
    videoPaths = []
    baseName = video.camera.trapgroup.tag + '_' + stringify_timestamp(videoImages[0].corrected_timestamp, True)
    for label in videoLabels:
            
            if (label.id==GLOBALS.nothing_id) and not include_empties:
                continue
            
            individuals = [None]
            
            if individual_sorted:
                individuals = db.session.query(Individual)\
                                    .join(Detection,Individual.detections)\
                                    .join(Image)\
                                    .filter(Image.id.in_(video_images_ids))\
                                    .filter(Individual.tasks.contains(task))\
                                    .filter(Individual.species==label.description)\
                                    .filter(Individual.active==True)\
                                    .distinct().all()
                if len(individuals)==0: individuals = [None]
            
            for individual in individuals:
                if label.id in requested_labels:
                    videoPath = ''
                    if species_sorted:  videoPath += '/' + label.description.replace('/','_').replace('\\','_')
                    if individual and individual_sorted: videoPath += '/' + individual.name
                    if flat_structure:
                        filename = baseName
                        for videoLabel in videoLabels: filename += '_' + videoLabel.description.replace(' ','_').replace('/','_').replace('\\','_')
                        filename += '_' + str(video.id) + '.mp4'
                        videoPath += '/' + filename
                    else:
                        startPoint = 1
                        if splitPath[1]==task.survey.name: startPoint=2
                        for split in splitPath[startPoint:]: videoPath += '/' + split
                        videoPath += '/' +video.filename.rsplit('.', 1)[0] + '.mp4'
                    videoPaths.append(videoPath)

    return list(set(videoPaths)), [label.description for label in videoLabels], [tag.description for tag in videoTags]

# def prepare_exif_image(image_id,task_id,species_sorted,flat_structure,individual_sorted,surveyName,labels):
#     '''
#     Processes a single image for exif download by downloading the file locally, editing its metadata (without opening it) and saving it 
#     to a Downloads folder in the user's bucket. Labels are saved in the user comment exif data, xpkeyword data and the IPTC keywords.
    
#         Parameters:
#             image_id (int): The image to process
#             task_id (int): The task whose labels must be used
#             species_sorted (bool): Whether the dataset should be sorted into species folders
#             flat_structure (bool): Whether the folder structure should be flattened
#             individual_sorted (bool): Wether to sort the data by individuals
#             surveyName (str): The name of the survey associated with the image
#             labels (list): The requested label ids
#     '''
#     try:
#         image = db.session.query(Image).get(image_id)
#         sourceKey = image.camera.path + '/' + image.filename
#         task = db.session.query(Task).get(task_id)

#         imagePaths, imageLabels, imageTags = get_image_paths_and_labels(image,task,individual_sorted,species_sorted,flat_structure,labels)

#         destinationKeys = []
#         for path in imagePaths:
#             destinationKeys.append(task.survey.user.folder + '/Downloads/' + surveyName + '/' + task.name + '/' + path)

#         # with tempfile.NamedTemporaryFile(delete=True, suffix='.JPG') as temp_file:
#         # GLOBALS.s3client.download_file(Bucket=bucket, Key=sourceKey, Filename=temp_file.name)

#         s3_response_object = GLOBALS.s3client.get_object(Bucket=Config.BUCKET,Key=sourceKey)
#         imageData = s3_response_object['Body'].read()

#         exifData = b'ASCII\x00\x00\x00'
#         xpKeywordData = ''
#         # IPTCData = []
#         imageLabels.extend(imageTags)
#         for label in imageLabels:
#             xpKeywordData += label
#             exifData += label.encode()
#             # IPTCData.append(label.encode())
#             if label != imageLabels[-1]:
#                 xpKeywordData += ','
#                 exifData += b', '

#         # EXIF
#         try:
#             try:
#                 exif_dict = piexif.load(imageData)
#                 exif_bytes = piexif.dump(exif_dict)
#             except:
#                 # If exif data is corrupt, then just overwite it entirely
#                 exif_dict={'0th':{},'Exif':{}}
#             exif_dict['Exif'][37510] = exifData #write the data to the user comment exif data
#             exif_dict['Exif'][36867] = image.corrected_timestamp.strftime("%Y/%m/%d %H:%M:%S").encode() #created on 
#             exif_dict['Exif'][36868] = image.corrected_timestamp.strftime("%Y/%m/%d %H:%M:%S").encode()
#             exif_dict['0th'][40094] = xpKeywordData.encode('utf-16')
#             exif_bytes = piexif.dump(exif_dict)
#             output=io.BytesIO()
#             piexif.insert(exif_bytes,imageData,output) #insert new exif data without opening & re-saving image
#         except:
#             # Rather ensure that the image is there, without exif data than the opposite.
#             pass

#         # # IPTC
#         # try:
#         #     info = IPTCInfo(temp_file.name)
#         #     info['keywords'] = IPTCData
#         #     info.save()
#         # except:
#         #     # Rather ensure image is there
#         #     pass

#         for destinationKey in destinationKeys:
#             GLOBALS.s3client.put_object(Body=output,Bucket=Config.BUCKET,Key=destinationKey)
#             # GLOBALS.s3client.upload_file(Filename=temp_file.name, Bucket=bucket, Key=destinationKey)

#     except Exception:
#         app.logger.info(' ')
#         app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
#         app.logger.info(traceback.format_exc())
#         app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
#         app.logger.info(' ')

#     finally:
#         db.session.remove()

#     return True

# @celery.task(bind=True,max_retries=5)
# def prepare_exif_batch(self,image_ids,task_id,species_sorted,flat_structure,individual_sorted,surveyName,labels):
#     ''' Prepares a batch of exif images, allowing for parallelisation across instances. 
    
#         Parameters:
#             image_ids (list): The batch of images to process
#             task_id (int): The task whose labels must be used
#             species_sorted (bool): Whether the dataset should be sorted into species folders
#             flat_structure (bool): Whether the folder structure should be flattened
#             individual_sorted (bool): Whether the images should be sorted by individuals
#             surveyName (str): The name of the survey associated with the image
#             labels (list): The requested label ids
#     '''

#     try:
#         for image_id in image_ids:
#             prepare_exif_image(image_id,task_id,species_sorted,flat_structure,individual_sorted,surveyName,labels)

#     except Exception as exc:
#         app.logger.info(' ')
#         app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
#         app.logger.info(traceback.format_exc())
#         app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
#         app.logger.info(' ')
#         self.retry(exc=exc, countdown= retryTime(self.request.retries))

#     finally:
#         db.session.remove()

#     return True

# @celery.task(bind=True,max_retries=5,ignore_result=True)
# def prepare_exif(self,task_id,species,species_sorted,flat_structure,individual_sorted):
#     '''
#     Celery task for preparing an exif dataset download. Triggers the threaded processing of images.

#         Parameters:
#             task_id (int): The task for which exif dataset is needed
#             species (list): The species for download
#             species_sorted (bool): Whether the dataset should be sorted into species folders
#             flat_structure (bool): Whether the folder structure should be flattened
#             individual_sorted (bool): Whether the data should be sorted by identified individuals (where possible)
#     '''
#     try:
#         app.logger.info('prepare_exif started for task {}'.format(task_id))
#         task = db.session.query(Task).get(task_id)
#         task.survey.status = 'Processing'
#         task.download_available = False
#         db.session.commit()
#         surveyName = task.survey.name

#         # Delete previous
#         s3 = boto3.resource('s3')
#         bucketObject = s3.Bucket(Config.BUCKET)
#         bucketObject.objects.filter(Prefix=task.survey.user.folder+'/Downloads/'+surveyName+'/'+task.name+'/').delete()
        
#         if '0' in species:
#             labels = db.session.query(Label).filter(Label.task_id==task_id).distinct().all()
#             labels.append(db.session.query(Label).get(GLOBALS.vhl_id))
#         else:
#             labels = db.session.query(Label).filter(Label.id.in_(species)).distinct().all()

#         images = db.session.query(Image)\
#                         .join(Detection)\
#                         .join(Labelgroup)\
#                         .join(Label,Labelgroup.labels)\
#                         .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
#                         .filter(Detection.static==False)\
#                         .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
#                         .filter(Labelgroup.task_id==task.id)\
#                         .filter(Label.id.in_([r.id for r in labels]))\
#                         .distinct().all()

#         task.survey.images_processing=len(images)
#         db.session.commit()

#         results = []
#         for batch in chunker(images,5000):
#             results.append(prepare_exif_batch.apply_async(kwargs={  'image_ids':[r.id for r in batch],
#                                                                     'task_id':task_id,
#                                                                     'species_sorted':species_sorted,
#                                                                     'flat_structure':flat_structure,
#                                                                     'individual_sorted':individual_sorted,
#                                                                     'surveyName': surveyName,
#                                                                     'labels': [r.id for r in labels]},queue='parallel'))

#         #Wait for processing to complete
#         # Using locking here as a workaround. Looks like celery result fetching is not threadsafe.
#         # See https://github.com/celery/celery/issues/4480
#         db.session.remove()
#         GLOBALS.lock.acquire()
#         with allow_join_result():
#             for result in results:
#                 try:
#                     result.get()
#                 except Exception:
#                     app.logger.info(' ')
#                     app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
#                     app.logger.info(traceback.format_exc())
#                     app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
#                     app.logger.info(' ')
#                 result.forget()
#         GLOBALS.lock.release()

#         task = db.session.query(Task).get(task_id)
#         task.survey.status = 'Ready'
#         task.download_available = True
#         db.session.commit()

#     except Exception as exc:
#         app.logger.info(' ')
#         app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
#         app.logger.info(traceback.format_exc())
#         app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
#         app.logger.info(' ')
#         self.retry(exc=exc, countdown= retryTime(self.request.retries))

#     finally:
#         db.session.remove()

#     return True

@celery.task(bind=True,max_retries=5,ignore_result=True)
def generate_training_csv(self,tasks,destBucket,min_area,include_empties=False,crop_images=True):
    '''
    Generates a csv file for classification training purposes.

        Parameters:
            tasks (list): A list of task IDs for which a file should be created
            destBucket (str): The destination bucket for image crops
            min_area (float): The minimum normalised size of detection crops
    '''

    try:
        organisations = db.session.query(Organisation).join(Survey).join(Task).filter(Task.id.in_(tasks)).distinct().all()
        if len(organisations) == 1:
            organisation = organisations[0].name
        else:
            organisation = 'Multiple'
        
        key = 'classification_ds/'+randomString()+organisation+'_classification_ds.csv'

        for task_id in tasks:
            task = db.session.query(Task).get(task_id)

            # we want to include all child-level labels
            labels = [r[0] for r in db.session.query(Label.id).filter(Label.task_id==task_id).filter(~Label.children.any()).all()]
            if include_empties: labels.append(GLOBALS.nothing_id)

            df = pd.read_sql(db.session.query(\
                        Detection.id.label('detection_id'),\
                        Detection.score.label('confidence'),\
                        Detection.left.label('left'),\
                        Detection.right.label('right'),\
                        Detection.top.label('top'),\
                        Detection.bottom.label('bottom'),\
                        Cluster.id.label('cluster_id'),\
                        Image.id.label('image_id'),\
                        Image.filename.label('filename'),\
                        Image.hash.label('hash'),\
                        Camera.path.label('dirpath'),\
                        Trapgroup.tag.label('location'),\
                        Survey.name.label('dataset'),\
                        Label.description.label('label'))\
                        .join(Image,Detection.image_id==Image.id)\
                        .join(Cluster,Image.clusters)\
                        .join(Camera,Camera.id==Image.camera_id)\
                        .join(Trapgroup,Trapgroup.id==Camera.trapgroup_id)\
                        .join(Survey,Survey.id==Trapgroup.survey_id)\
                        .join(Labelgroup,Labelgroup.detection_id==Detection.id)\
                        .join(Label,Labelgroup.labels)\
                        .filter(Labelgroup.task_id==task_id)\
                        .filter(Cluster.task_id==task_id)\
                        .filter(Label.id.in_(labels))\
                        .filter(((Detection.right-Detection.left)*(Detection.bottom-Detection.top)) > min_area)\
                        .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                        .filter(Detection.static == False)\
                        .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                        .statement,db.session.bind)

            # Drop multi labels
            df = df.drop_duplicates(subset=['detection_id'], keep=False)

            if len(df):
                # # Build crop names
                # df['path'] = df.apply(lambda x: x.dirpath+'/'+x.filename[:-4]+'_'+str(x.detection_id)+'.jpg', axis=1)

                # Build full paths
                df['path'] = df.apply(lambda x: x.dirpath+'/'+x.filename, axis=1)

                # Set the dataset_class to the label value
                df['dataset_class'] = df['label']

                # Check if you need to crop the images
                # if crop_images:
                #     try:              
                #         key = df.iloc[0]['path']
                #         check = GLOBALS.s3client.head_object(Bucket=destBucket,Key=key)
                    
                #     except:
                #         # Crop does not exist - must crop the images
                #         crop_survey_images.apply_async(kwargs={'task_id':task_id,'min_area':min_area,'destBucket':destBucket},queue='parallel')
                #         task.survey.images_processing = len(df)
                #         # task.survey.status='Processing'
                #         db.session.commit()

                # Int locations cause an issue
                df['location'] = df.apply(lambda x: str(x.location), axis=1)
                df['location'] = df.apply(lambda x: 'site_'+x.location if x.location.isdigit() else x.location, axis=1)

                # Order columns and remove superfluous ones
                df = df[['detection_id','left','right','top','bottom','cluster_id','path','hash','dataset','location','dataset_class','confidence','label']]

                # convert all labels to lower case
                df['label'] = df.apply(lambda x: x.label.lower().strip(), axis=1)
                df['dataset_class'] = df.apply(lambda x: x.dataset_class.lower().strip(), axis=1)

                # Add to output
                # if outputDF is not None:
                #     outputDF = pd.concat([outputDF, df], ignore_index=True)
                #     outputDF.fillna(0, inplace=True)
                # else:
                #     outputDF = df

                df.to_csv(key, index=False, mode='a', header=not os.path.exists(key))

        # Write output to S3
        GLOBALS.s3client.upload_file(Filename=key, Bucket=destBucket, Key=key)
        os.remove(key)

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        self.retry(exc=exc, countdown= retryTime(self.request.retries))

    finally:
        db.session.remove()

    return key

@celery.task(bind=True,max_retries=5,ignore_result=True)
def crop_survey_images(self,task_id,min_area,destBucket,include_empties=False):
    '''
    Helper task for generate_training_csv that allows the image-cropping process to be parallelised.

    Parameters:
        task_id (int): The task you are using to crop the associated survey
        min_area (float): The minimum detection area to crop
        destBucket (str): The bucket where the cropped images must be stored
    '''

    try:
        task = db.session.query(Task).get(task_id)

        # we want to include all child-level labels
        labels = db.session.query(Label).filter(Label.task_id==task_id).filter(~Label.children.any()).all()
        labels = [r[0] for r in db.session.query(Label.id).filter(Label.task_id==task_id).filter(~Label.children.any()).all()]
        if include_empties: labels.append(GLOBALS.nothing_id)

        df = pd.read_sql(db.session.query(\
                    Detection.id.label('detection_id'),\
                    Detection.score.label('confidence'),\
                    Image.id.label('image_id'),\
                    Image.filename.label('filename'),\
                    Camera.path.label('dirpath'),\
                    Trapgroup.tag.label('location'),\
                    Survey.name.label('dataset'),\
                    Label.description.label('label'))\
                    .join(Image,Detection.image_id==Image.id)\
                    .join(Camera,Camera.id==Image.camera_id)\
                    .join(Trapgroup,Trapgroup.id==Camera.trapgroup_id)\
                    .join(Survey,Survey.id==Trapgroup.survey_id)\
                    .join(Labelgroup,Labelgroup.detection_id==Detection.id)\
                    .join(Label,Labelgroup.labels)\
                    .filter(Labelgroup.task_id==task_id)\
                    .filter(Label.id.in_(labels))\
                    .filter(((Detection.right-Detection.left)*(Detection.bottom-Detection.top)) > min_area)\
                    .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                    .filter(Detection.static == False)\
                    .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                    .statement,db.session.bind)

        # Drop detections
        df = df.drop_duplicates(subset=['detection_id'], keep=False)

        results = []
        for chunk in chunker(df['image_id'].unique(),10000):
            results.append(batch_crops.apply_async(kwargs={'image_ids':[int(r) for r in chunk],'source':None,'min_area':min_area,'destBucket':destBucket,'external':False,'update_image_info':False},queue='default'))

        # Using locking here as a workaround. Looks like celery result fetching is not threadsafe.
        # See https://github.com/celery/celery/issues/4480
        GLOBALS.lock.acquire()
        with allow_join_result():
            for result in results:
                try:
                    result.get()
                except Exception:
                    app.logger.info(' ')
                    app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
                    app.logger.info(traceback.format_exc())
                    app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
                    app.logger.info(' ')
                result.forget()
        GLOBALS.lock.release()

        # task.survey.images_processing = 0
        # task.survey.status='Ready'
        db.session.commit()

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        self.retry(exc=exc, countdown= retryTime(self.request.retries))

    finally:
        db.session.remove()

    return True

@celery.task(bind=True,max_retries=5,ignore_result=True)
def generate_classification_ds(self,sourceBucket):
    '''Searches for all the classification_ds files in the source bucket, combines them, and uploades the result to the bucket.'''

    try:
        outputDF = None
        folders,filenames = list_all(sourceBucket,'classification_ds/')
        for filename in filenames:
            if filename != 'classification_ds.csv':
                with tempfile.NamedTemporaryFile(delete=True, suffix='.csv') as temp_file:
                    GLOBALS.s3client.download_file(Bucket=sourceBucket, Key='classification_ds/'+filename, Filename=temp_file.name)
                    df = pd.read_csv(temp_file.name)

                # Add to output
                if outputDF is not None:
                    outputDF = pd.concat([outputDF, df], ignore_index=True)
                    outputDF.fillna(0, inplace=True)
                else:
                    outputDF = df

        # Write output to S3
        with tempfile.NamedTemporaryFile(delete=True, suffix='.csv') as temp_file:
            outputDF.to_csv(temp_file.name,index=False)
            GLOBALS.s3client.put_object(Bucket=sourceBucket,Key='classification_ds/classification_ds.csv',Body=temp_file)

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        self.retry(exc=exc, countdown= retryTime(self.request.retries))

    finally:
        db.session.remove()

    return True


@celery.task(bind=True,max_retries=5,ignore_result=True)
def generate_label_spec(self,sourceBucket,translations):
    '''
    Generates a label spec file for classifier training purposes based on the supplied translations. 
    Saves the resulting file in the source bucket.
    '''

    try:
        with tempfile.NamedTemporaryFile(delete=True, suffix='.csv') as temp_file:
            GLOBALS.s3client.download_file(Bucket=sourceBucket, Key='classification_ds/classification_ds.csv', Filename=temp_file.name)
            df = pd.read_csv(temp_file.name)

        #Generate dictionary of dataset labels
        dataset_labels = {}
        datasets = df['dataset'].unique()
        for dataset in datasets:
            dataset_labels[dataset] = df[df['dataset']==dataset]['label'].unique()

        # Generate Label Spec
        label_spec = {}
        for label in translations:
            label_spec[label] = {'dataset_labels':{}}
            for dataset in datasets:
                label_spec[label]['dataset_labels'][dataset] = []
                for translation in translations[label]:
                    if translation in dataset_labels[dataset]:
                        label_spec[label]['dataset_labels'][dataset].append(translation)

        with io.BytesIO() as f:
            f.write(json.dumps(label_spec).encode())
            f.seek(0)
            GLOBALS.s3client.put_object(Bucket=sourceBucket,Key='label_spec.json',Body=f)

        # Generate accompanying label index
        index = 0
        label_index = {}
        for label in translations:
            label_index[index] = label
            index += 1

        with io.BytesIO() as f:
            f.write(json.dumps(label_index).encode())
            f.seek(0)
            GLOBALS.s3client.put_object(Bucket=sourceBucket,Key='label_index.json',Body=f)

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        self.retry(exc=exc, countdown= retryTime(self.request.retries))

    finally:
        db.session.remove()

    return True

@celery.task(bind=True,max_retries=1,ignore_result=True)
def generate_coco(self,task_id,user_name,download_id):
    '''Generates a COCO export of a task.'''

    try:
        task = db.session.query(Task).get(task_id)
        # fileName = task.survey.user.folder+'/docs/'+task.survey.user.username+'_'+task.survey.name+'_'+task.name+'.json'
        randomness = randomString()
        fileName = task.survey.organisation.folder+'/docs/'+task.survey.organisation.name+'_'+user_name+'_'+task.survey.name+'_'+task.name+'_'+randomness+'.json'

        # # Delete old file if exists
        # try:
        #     GLOBALS.s3client.delete_object(Bucket=Config.BUCKET, Key=fileName)
        # except:
        #     pass

        info = {
            "version" : 1,
            "description" : task.survey.description,
            "year" : datetime.utcnow().year,
            "contributor" : task.survey.organisation.name,
            "date_created" : str(datetime.utcnow())
        }
        
        images = []
        for cluster in task.clusters:
            for image in cluster.images:
                images.append({
                    "id" : str(image.id),
                    "file_name" : '/'.join(image.camera.path.split('/')[1:])+'/'+image.filename,
                    "datetime": stringify_timestamp(image.corrected_timestamp),
                    "seq_id": str(cluster.id),
                    "seq_num_frames": len(cluster.images[:]),
                    "location": image.camera.trapgroup.tag,
                    "corrupt": False
                })

        labels = [db.session.query(Label).get(GLOBALS.nothing_id)]
        labels.extend(db.session.query(Label)\
                        .join(Labelgroup,Label.labelgroups)\
                        .filter(Labelgroup.task_id==task_id)\
                        .filter(Label.id!=GLOBALS.nothing_id)\
                        .distinct().all())

        categories = []   
        for label in labels:
            categories.append({
                "id" : label.id,
                "name" : label.description  
            })

        labelgroups = db.session.query(Labelgroup)\
                            .join(Detection)\
                            .filter(Labelgroup.task==task)\
                            .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                            .filter(Detection.static==False)\
                            .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                            .distinct().all()

        annotations = []
        for labelgroup in labelgroups:
            bbox = [labelgroup.detection.left,
                    labelgroup.detection.top,
                    labelgroup.detection.right-labelgroup.detection.left,
                    labelgroup.detection.bottom-labelgroup.detection.top]
            if labelgroup.labels:
                for label in labelgroup.labels:
                    annotations.append({
                        "id" : str(labelgroup.detection.id),
                        "image_id" : str(labelgroup.detection.image_id),  
                        "category_id" : label.id,
                        "bbox": bbox,
                        "sequence_level_annotation" : False
                    })
            else:
                annotations.append({
                    "id" : str(labelgroup.detection.id),
                    "image_id" : str(labelgroup.detection.image_id),  
                    "category_id" : GLOBALS.nothing_id,
                    "bbox": bbox,
                    "sequence_level_annotation" : False
                })
        
        output = {
            "info" : info,
            "images" : images,
            "categories" : categories,
            "annotations" : annotations
        }

        # Write new file to S3 for fetching
        with tempfile.NamedTemporaryFile(delete=True, suffix='.json') as temp_file:
            with open(temp_file.name, 'w') as f:
                json.dump(output, f)
            GLOBALS.s3client.put_object(Bucket=Config.BUCKET,Key=fileName,Body=temp_file)

        # Schedule deletion
        # deleteFile.apply_async(kwargs={'fileName': fileName}, countdown=3600)

        download_request = db.session.query(DownloadRequest).get(download_id)
        if download_request:
            download_request.status = 'Available'
            download_request.timestamp = datetime.now() + timedelta(days=7)
            download_request.name = randomness
            db.session.commit()

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        self.retry(exc=exc, countdown= retryTime(self.request.retries))

    finally:
        db.session.remove()

    return True

@celery.task(bind=True,soft_time_limit=82800)
def calculate_results_summary(self, task_ids, baseUnit, sites, groups, startDate, endDate, user_id, trapUnit, timeToIndependence, timeToIndependenceUnit, normaliseBySite):
    ''' Calculates the results summary '''
    try:
        summary = {}

        if task_ids:
            if task_ids[0] == '0':
                tasks = surveyPermissionsSQ(db.session.query(Task.id, Task.survey_id).join(Survey).filter(Task.name != 'default').filter(~Task.name.contains('_o_l_d_')).filter(~Task.name.contains('_copying')).group_by(Task.survey_id).order_by(Task.id), user_id, 'read').distinct().all()
            else:
                tasks = surveyPermissionsSQ(db.session.query(Task.id, Task.survey_id).join(Survey).filter(Task.id.in_(task_ids)), user_id, 'read').distinct().all()

            task_ids = [r[0] for r in tasks]
            survey_ids = list(set([r[1] for r in tasks]))

            # Data Summary 
            # Get the total detection counts, total image counts and total cluster counts and the first and last date of the data 
            summaryQuery = db.session.query(
                func.count(distinct(Cluster.id)),
                func.count(distinct(Image.id)),
                func.count(distinct(Detection.id)),
                func.min(Image.corrected_timestamp),
                func.max(Image.corrected_timestamp),
            )\
            .join(Image, Cluster.images)\
            .join(Detection)\
            .join(Labelgroup)\
            .join(Camera)\
            .join(Trapgroup)\
            .outerjoin(Sitegroup, Trapgroup.sitegroups)\
            .filter(Cluster.task_id.in_(task_ids))\
            .filter(Labelgroup.task_id.in_(task_ids))\
            .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
            .filter(Detection.static==False)\
            .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))

            if startDate: summaryQuery = summaryQuery.filter(Image.corrected_timestamp >= startDate)

            if endDate: summaryQuery = summaryQuery.filter(Image.corrected_timestamp <= endDate)

            if sites != '0' and sites != '-1' and groups != '0' and groups != '-1':
                summaryQuery = summaryQuery.filter(or_(Trapgroup.id.in_(sites),Sitegroup.id.in_(groups)))
            elif sites != '0' and sites != '-1':
                summaryQuery = summaryQuery.filter(Trapgroup.id.in_(sites))
            elif groups != '0' and groups != '-1':
                summaryQuery = summaryQuery.filter(Sitegroup.id.in_(groups))

            summaryTotals = summaryQuery.all()
            
            vhl = db.session.query(Label).get(GLOBALS.vhl_id)
            label_list = [GLOBALS.vhl_id,GLOBALS.nothing_id,GLOBALS.knocked_id]
            for task_id in task_ids:
                label_list.extend(getChildList(vhl,int(task_id)))
            summaryQuery = summaryQuery.filter(~Labelgroup.labels.any(Label.id.in_(label_list))).filter(Labelgroup.labels.any())

            summaryAnimalTotals = summaryQuery.all()   

            summary_counts = {
                'Total Clusters': {
                    'value': summaryTotals[0][0],
                    'description': 'The total number of clusters in your surveys that have valid sightings.'
                },
                'Total Animal Clusters': {
                    'value': summaryAnimalTotals[0][0],
                    'description': 'The total number of clusters in the data that contain animal sightings (excluding Vehicles/Humans/Livestock, Nothing or Knocked sightings).'
                },	
                'Total Images': {
                    'value': summaryTotals[0][1],
                    'description': 'The total number of images in your surveys that have valid sightings.'
                },
                'Total Animal Images': {
                    'value': summaryAnimalTotals[0][1],
                    'description': 'The total number of images in the data that contain animal sightings (excluding Vehicles/Humans/Livestock, Nothing or Knocked sightings).'
                },
                'Total Sightings': {
                    'value': summaryTotals[0][2],
                    'description': 'The overall count of valid sightings in your surveys.'
                },
                'Total Animal Sightings': {
                    'value': summaryAnimalTotals[0][2],
                    'description': 'The total number of sightings of animals (excluding Vehicles/Humans/Livestock, Nothing or Knocked sightings) in your surveys.'
                },
                'First Date': {
                    'value': summaryTotals[0][3].strftime('%Y-%m-%d %H:%M:%S') if summaryTotals[0][3] else None,
                    'description': 'The date and time of the first recorded sighting in your surveys.'
                },
                'Last Date': {
                    'value': summaryTotals[0][4].strftime('%Y-%m-%d %H:%M:%S') if summaryTotals[0][4] else None,
                    'description': 'The date and time of the most recent recorded sighting in your surveys.'
                }
            }

            summary['summary_counts'] = summary_counts

            # General query for all results
            if baseUnit == '1' or baseUnit == '4': # Image
                baseQuery = db.session.query(
                                Image.id,
                                Image.corrected_timestamp,
                                Label.id,
                                Label.description,
                                Camera.id,
                                Camera.path,
                                Trapgroup.id,
                                Trapgroup.tag,
                                Trapgroup.latitude,
                                Trapgroup.longitude,
                                Trapgroup.altitude,
                                Sitegroup.id,
                                Sitegroup.name,
                                Cameragroup.id,
                                Cameragroup.name
                            )\
                            .join(Detection)\
                            .join(Labelgroup)\
                            .join(Label,Labelgroup.labels)\
                            .join(Camera)\
                            .join(Cameragroup)\
                            .join(Trapgroup)\
                            .outerjoin(Sitegroup, Trapgroup.sitegroups)\
                            .filter(Labelgroup.task_id.in_(task_ids))\
                            .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                            .filter(Detection.static==False)\
                            .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                            .filter(Trapgroup.survey_id.in_(survey_ids))

            elif baseUnit == '2': # Cluster
                baseQuery = db.session.query(
                                Cluster.id,
                                Image.corrected_timestamp,
                                Label.id,
                                Label.description,
                                Camera.id,
                                Camera.path,
                                Trapgroup.id,
                                Trapgroup.tag,
                                Trapgroup.latitude,
                                Trapgroup.longitude,
                                Trapgroup.altitude,
                                Sitegroup.id,
                                Sitegroup.name,
                                Cameragroup.id,
                                Cameragroup.name
                            )\
                            .join(Image,Cluster.images)\
                            .join(Detection)\
                            .join(Labelgroup)\
                            .join(Label,Labelgroup.labels)\
                            .join(Camera)\
                            .join(Cameragroup)\
                            .join(Trapgroup)\
                            .outerjoin(Sitegroup, Trapgroup.sitegroups)\
                            .filter(Labelgroup.task_id.in_(task_ids))\
                            .filter(Cluster.task_id.in_(task_ids))\
                            .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                            .filter(Detection.static==False)\
                            .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                            .filter(Trapgroup.survey_id.in_(survey_ids))

            elif baseUnit == '3':  # Detection
                baseQuery = db.session.query(
                                Detection.id,
                                Image.corrected_timestamp,
                                Label.id,
                                Label.description,
                                Camera.id,
                                Camera.path,
                                Trapgroup.id,
                                Trapgroup.tag,
                                Trapgroup.latitude,
                                Trapgroup.longitude,
                                Trapgroup.altitude,
                                Sitegroup.id,
                                Sitegroup.name,
                                Cameragroup.id,
                                Cameragroup.name
                            )\
                            .join(Image, Detection.image_id==Image.id)\
                            .join(Camera)\
                            .join(Cameragroup)\
                            .join(Trapgroup)\
                            .join(Labelgroup)\
                            .join(Label,Labelgroup.labels)\
                            .outerjoin(Sitegroup, Trapgroup.sitegroups)\
                            .filter(Labelgroup.task_id.in_(task_ids))\
                            .filter(or_(and_(Detection.source==model,Detection.score>Config.DETECTOR_THRESHOLDS[model]) for model in Config.DETECTOR_THRESHOLDS))\
                            .filter(Detection.static==False)\
                            .filter(~Detection.status.in_(Config.DET_IGNORE_STATUSES))\
                            .filter(Trapgroup.survey_id.in_(survey_ids))
                
            if startDate: baseQuery = baseQuery.filter(Image.corrected_timestamp >= startDate)

            if endDate: baseQuery = baseQuery.filter(Image.corrected_timestamp <= endDate)

            if sites != '0' and sites != '-1' and groups != '0' and groups != '-1':
                baseQuery = baseQuery.filter(or_(Trapgroup.id.in_(sites),Sitegroup.id.in_(groups)))
            elif sites != '0' and sites != '-1':
                baseQuery = baseQuery.filter(Trapgroup.id.in_(sites))
            elif groups != '0' and groups != '-1':
                baseQuery = baseQuery.filter(Sitegroup.id.in_(groups))

            if baseUnit == '1' or baseUnit == '4':
                baseQuery = baseQuery.group_by(Image.id, Label.id, Camera.id, Trapgroup.id, Sitegroup.id).all()
            elif baseUnit == '2':
                baseQuery = baseQuery.group_by(Cluster.id, Label.id, Camera.id, Trapgroup.id, Sitegroup.id).all()
            elif baseUnit == '3':
                baseQuery = baseQuery.group_by(Detection.id, Label.id, Camera.id, Trapgroup.id, Sitegroup.id).all()

            df = pd.DataFrame(baseQuery, columns=['id','timestamp','label_id','species', 'camera_id', 'path', 'site_id', 'name', 'latitude', 'longitude', 'altitude', 'group_id', 'group_name', 'cameragroup_id', 'cameragroup_name'])

            # Unique rows
            df = df.drop_duplicates()

            # Group rows by species and site and check if the time between rows is less than the timeToIndependence and remove rows if so
            if timeToIndependence:
                if timeToIndependenceUnit == 's':
                    timeToIndependence = int(timeToIndependence)
                elif timeToIndependenceUnit == 'm':
                    timeToIndependence = int(timeToIndependence) * 60
                elif timeToIndependenceUnit == 'h':
                    timeToIndependence = int(timeToIndependence) * 3600
                timeToIndependence = timedelta(seconds=timeToIndependence)
                
                df = df.sort_values(by=['species','name', 'latitude', 'longitude','timestamp'])	
                df['timedelta'] = df.groupby(['species','name','latitude','longitude'])['timestamp'].diff()
                df['timedelta'] = df['timedelta'].fillna(timedelta(seconds=9999999))
                df = df[df['timedelta'] >= timeToIndependence]
                df = df.drop(columns=['timedelta'])

            if len(df) > 0:
                # Species count
                label_list.append(GLOBALS.unknown_id)
                species_df = df[~df['label_id'].isin(label_list)].copy()
                species_df = species_df[['id','label_id','species']]
                species_df.drop_duplicates(subset=['id','species'], inplace=True)
                species_df['count'] = 1
                species_count = species_df.groupby('species').sum().reset_index().drop('id', axis=1)

                # get total count
                total_count = species_count['count'].sum()
            else:
                species_count = pd.DataFrame(columns=['species','count'])
                total_count = 0

            # Diversity Indexes
            shannon_index = 0
            simpsons_index = 0
            hill_number = 0
            pielous_evenness = 0
            species_richness = 0

            if len(species_count) > 0:
                # Convert species_count count column to a list 
                species_count_list = species_count['count'].tolist()
                total_count = sum(species_count_list)

                # Calculate Shannon-Wiener Index
                shannon_index = -sum([(count/total_count)*math.log(count/total_count) for count in species_count_list if count != 0])

                # Calculate Simpson's Index
                simpsons_index = 1 - sum([(count/total_count)**2 for count in species_count_list if count != 0])

                # Calculate Hill Number
                hill_number = 1/sum([(count/total_count)**2 for count in species_count_list if count != 0])

                # Calculate Pielou's Evenness
                pielous_evenness = shannon_index/math.log(len(species_count_list))

                # Calculate Species Richness
                species_richness = len(species_count_list)

            summary_indexes = {
                'Shannon-Wiener Index': {
                    'value': shannon_index,
                    'description': 'The Shannon-Wiener index is a widely used measure of biodiversity that takes into account both species richness and evenness in a community. It provides a comprehensive assessment of diversity, where higher values indicate greater diversity. The Shannon-Wiener index typically ranges from 1.5 to 3.5.',
                },
                "Simpson's Index of Diversity": {
                    'value': simpsons_index,
                    'description': "Simpson's index of Diversity (1-D) is a measure of biodiversity that quantifies the probability that two individuals randomly selected from the community belong to different species. It ranges from 0 to 1, where 1 represents infinite diversity (all species equally abundant) and 0 represents minimum diversity (one species dominates the community). Higher values indicate higher diversity.",
                },
                'Hill Number (q=2)': {
                    'value': hill_number,
                    'description': 'The Hill number is a family of diversity indices that offers a unified way to measure biodiversity at different orders. Higher Hill numbers indicate higher diversity. Hill Number 0 represents species richness, Hill Number 1 is equivalent to the Shannon-Wiener index, and Hill Number 2 is equivalent to the inverse Simpson’s index (1/D).',
                },
                'Pielou\'s Evenness': {
                    'value': pielous_evenness,
                    'description': "Pielou's evenness is a measure of how evenly individuals are distributed among different species in a community. It ranges from 0 to 1, where higher values suggest a more even distribution of individuals among species.",
                },
                'Species Richness': {
                    'value': species_richness,
                    'description': 'Species richness is a simple measure of biodiversity that counts the number of different species present in a community. The higher the value, the more species there are in the community.',
                }
            }

            summary['summary_indexes'] = summary_indexes

            # Camera Trap effort
            if trapUnit == '0':  # Sites
                baseTrapQuery = db.session.query(
                    Image.id,
                    Image.corrected_timestamp,
                    Trapgroup.id,
                    Trapgroup.tag,
                    Trapgroup.latitude,
                    Trapgroup.longitude,
                    Trapgroup.altitude,
                ).join(Camera, Image.camera_id == Camera.id)\
                .join(Trapgroup)\
                .outerjoin(Sitegroup, Trapgroup.sitegroups)\
                .filter(Trapgroup.survey_id.in_(survey_ids))\
                .filter(Image.corrected_timestamp != None)

                if startDate: baseTrapQuery = baseTrapQuery.filter(Image.corrected_timestamp >= startDate)

                if endDate: baseTrapQuery = baseTrapQuery.filter(Image.corrected_timestamp <= endDate)

                if sites != '0' and sites != '-1' and groups != '0' and groups != '-1':
                    baseTrapQuery = baseTrapQuery.filter(or_(Trapgroup.id.in_(sites),Sitegroup.id.in_(groups)))
                elif sites != '0' and sites != '-1':
                    baseTrapQuery = baseTrapQuery.filter(Trapgroup.id.in_(sites))
                elif groups != '0' and groups != '-1':
                    baseTrapQuery = baseTrapQuery.filter(Sitegroup.id.in_(groups))

                base_trap_df = pd.DataFrame(baseTrapQuery.all(), columns=['id','timestamp', 'site_id', 'name', 'latitude', 'longitude', 'altitude'])
                base_trap_df = base_trap_df.drop_duplicates()

                # Convert 'timestamp' column to datetime
                base_trap_df['timestamp'] = pd.to_datetime(base_trap_df['timestamp'])
                base_trap_df['date'] = base_trap_df['timestamp'].dt.date	

                # Effort Days
                # Group by 'site_id' and count the unique dates to get the number of days each site captured an image
                effort_days_df = base_trap_df.copy()
                effort_days_df = effort_days_df.groupby(['name', 'latitude', 'longitude'])['date'].nunique().reset_index()
                effort_days_df.rename(columns={'date': 'count'}, inplace=True)
                summary['effort_days'] = effort_days_df.sort_values(by='count', ascending=False).to_dict(orient='records')
                    
                # Active Days
                # Calculate which days each site was active for from a start date to an end date
                if startDate:
                    start_date = datetime.strptime(startDate.split(' ')[0], '%Y-%m-%d')
                else:
                    start_date = base_trap_df['timestamp'].min()

                if endDate:
                    end_date = datetime.strptime(endDate.split(' ')[0], '%Y-%m-%d')
                else:
                    end_date = base_trap_df['timestamp'].max()

                # Create a new id that will be assigned to sites that have the same site tag and coordinates
                trap_active_df = base_trap_df.copy()
                trap_active_df['site_id'] = trap_active_df.groupby(['name', 'latitude', 'longitude']).ngroup()
                trap_active_df = trap_active_df.drop_duplicates(subset=['id', 'timestamp', 'site_id'])

                # Add a count column to the trap_active_df DataFrame and set it to 1
                trap_active_df['count'] = 1

                # Group by 'site_id' and count the unique dates to get the number of days each site captured an image
                trap_active_df = trap_active_df.groupby(['site_id', 'name', 'latitude', 'longitude', 'altitude', 'date'])['count'].sum().reset_index()
                trap_active_dict = trap_active_df.sort_values(by='date', ascending=True).to_dict(orient='records')

                # Get the names of the sites for each site_id and convert to a list
                unit_names = trap_active_df.groupby('site_id')['name'].unique().explode().tolist()

                max_count = trap_active_df['count'].max()
                if np.isnan(max_count):
                    max_count = 0

                summary['active_days'] = {
                    'active_dict': trap_active_dict,
                    'start_date': start_date,
                    'end_date': end_date,
                    'unit_names': unit_names,
                    'max_count' : int(max_count)
                }
                
                # Unit Counts            
                trap_counts_df = df[['id', 'site_id', 'name', 'latitude', 'longitude', 'altitude']].copy()
                trap_counts_df = trap_counts_df.drop_duplicates(subset=['id', 'site_id'])
                trap_counts_df['count'] = 1

                # Create a new id that will be assigned to sites that have the same tag and lat and lng coordinates
                trap_counts_df['site_id'] = trap_counts_df.groupby(['name', 'latitude', 'longitude']).ngroup()        

                # Combine sites with same id and sum theur counts
                trap_counts_df = trap_counts_df.groupby(['site_id', 'name', 'latitude', 'longitude'])['count'].sum().reset_index()
                
                # Convert the DataFrame to a list of dictionaries
                summary['unit_counts'] = trap_counts_df.sort_values(by='count', ascending=False).to_dict(orient='records')

            elif trapUnit == '1':  # Cameras   
                # Base Camera Query
                baseCamQuery = db.session.query(
                    Image.id,
                    Image.corrected_timestamp,
                    Cameragroup.id,
                    Cameragroup.name,
                    Trapgroup.id,
                    Trapgroup.tag
                ).join(Camera, Image.camera_id == Camera.id)\
                .join(Cameragroup)\
                .join(Trapgroup)\
                .outerjoin(Sitegroup, Trapgroup.sitegroups)\
                .filter(Trapgroup.survey_id.in_(survey_ids))\
                .filter(Image.corrected_timestamp != None)

                if startDate: baseCamQuery = baseCamQuery.filter(Image.corrected_timestamp >= startDate)

                if endDate: baseCamQuery = baseCamQuery.filter(Image.corrected_timestamp <= endDate)

                if sites != '0' and sites != '-1' and groups != '0' and groups != '-1':
                    baseCamQuery = baseCamQuery.filter(or_(Trapgroup.id.in_(sites),Sitegroup.id.in_(groups)))
                elif sites != '0' and sites != '-1':
                    baseCamQuery = baseCamQuery.filter(Trapgroup.id.in_(sites))
                elif groups != '0' and groups != '-1':
                    baseCamQuery = baseCamQuery.filter(Sitegroup.id.in_(groups))
                
                base_cam_df = pd.DataFrame(baseCamQuery.distinct().all(), columns=['id','timestamp', 'cameragroup_id', 'name', 'site_id', 'site_tag'])
                base_cam_df = base_cam_df.drop_duplicates()

                # Convert 'timestamp' column to datetime
                base_cam_df['timestamp'] = pd.to_datetime(base_cam_df['timestamp'])
                base_cam_df['date'] = base_cam_df['timestamp'].dt.date

                # Effort Days
                # Group by name + site_tag and count the unique dates to get the number of days each camera captured an image and include the camera path
                effort_days_df = base_cam_df.copy()
                effort_days_df['name'] = effort_days_df['site_tag'] + '-' + effort_days_df['name']
                effort_days_df = effort_days_df.groupby(['name'])['date'].nunique().reset_index()
                effort_days_df.rename(columns={'date': 'count'}, inplace=True)

                summary['effort_days'] = effort_days_df.sort_values(by='count', ascending=False).to_dict(orient='records')

                # Active Days
                # Calculate which days each camera was active for from a start date to an end date
                if startDate:
                    start_date = datetime.strptime(startDate.split(' ')[0], '%Y-%m-%d')
                else:
                    start_date = base_cam_df['timestamp'].min()

                if endDate:
                    end_date = datetime.strptime(endDate.split(' ')[0], '%Y-%m-%d')
                else:
                    end_date = base_cam_df['timestamp'].max()

                # Create a new id that will be assigned to cameras that have the same name and site tag
                camera_active_df = base_cam_df.copy()
                camera_active_df['camera_id'] = camera_active_df.groupby(['site_tag', 'name']).ngroup()
                camera_active_df['name'] = camera_active_df['site_tag'] + '-' + camera_active_df['name']
                camera_active_df = camera_active_df.drop_duplicates(subset=['id', 'timestamp', 'camera_id']) 

                # Add a count column to the camera_active_df DataFrame and set it to 1
                camera_active_df['count'] = 1

                # Group by 'camera_id' and count the unique dates to get the number of days each camera captured an image
                camera_active_df = camera_active_df.groupby(['camera_id','name', 'date'])['count'].sum().reset_index()

                camera_active_dict = camera_active_df.sort_values(by='date', ascending=True).to_dict(orient='records')

                # Get the names of the cameras for each camera_id to a 1d list
                unit_names = camera_active_df.groupby('camera_id')['name'].unique().explode().tolist()

                max_count = camera_active_df['count'].max()

                summary['active_days'] = {
                    'active_dict': camera_active_dict,
                    'start_date': start_date,
                    'end_date': end_date,
                    'unit_names': unit_names,
                    'max_count' : int(max_count)
                }

                # Unit Counts    
                # Calculate how many counts each camera has capture for either clusters, images or detections
                # Dataframe with the number of counts for each camera
                camera_counts_df = df[['id', 'camera_id', 'path', 'site_id', 'name', 'cameragroup_id', 'cameragroup_name']].copy()
                camera_counts_df.rename(columns={'name': 'site_tag'}, inplace=True)
                camera_counts_df.rename(columns={'cameragroup_name': 'name'}, inplace=True)
                camera_counts_df = camera_counts_df.drop_duplicates(subset=['id', 'camera_id'])
                camera_counts_df['count'] = 1

                # Create a new id that will be assigned to cameras that have the same name and site tag
                camera_counts_df['name'] = camera_counts_df['site_tag'] + '-' + camera_counts_df['name']
                camera_counts_df['camera_id'] = camera_counts_df.groupby(['name']).ngroup()
                
                # Combine cameras with same id and sum theur counts
                camera_counts_df = camera_counts_df.groupby(['camera_id', 'name'])['count'].sum().reset_index()

                # Convert the DataFrame to a list of dictionaries
                summary['unit_counts'] = camera_counts_df.sort_values(by='count', ascending=False).to_dict(orient='records')

            elif trapUnit == '2':  # Sitegroups
                # Base Sitegroup Query
                baseGroupQuery = db.session.query(
                    Image.id,
                    Image.corrected_timestamp,
                    Sitegroup.id,
                    Sitegroup.name
                ).join(Camera, Image.camera_id == Camera.id)\
                .join(Trapgroup)\
                .join(Sitegroup, Trapgroup.sitegroups)\
                .filter(Trapgroup.survey_id.in_(survey_ids))\
                .filter(Image.corrected_timestamp != None)

                if startDate: baseGroupQuery = baseGroupQuery.filter(Image.corrected_timestamp >= startDate)

                if endDate: baseGroupQuery = baseGroupQuery.filter(Image.corrected_timestamp <= endDate)

                if sites != '0' and sites != '-1' and groups != '0' and groups != '-1':
                    baseGroupQuery = baseGroupQuery.filter(or_(Trapgroup.id.in_(sites),Sitegroup.id.in_(groups)))
                elif sites != '0' and sites != '-1':
                    baseGroupQuery = baseGroupQuery.filter(Trapgroup.id.in_(sites))
                elif groups != '0' and groups != '-1':
                    baseGroupQuery = baseGroupQuery.filter(Sitegroup.id.in_(groups))

                base_group_df = pd.DataFrame(baseGroupQuery.distinct().all(), columns=['id','timestamp','group_id', 'name'])
                base_group_df = base_group_df.drop_duplicates()

                # Convert 'timestamp' column to datetime
                base_group_df['timestamp'] = pd.to_datetime(base_group_df['timestamp'])
                base_group_df['date'] = base_group_df['timestamp'].dt.date	

                # Effort Days
                effort_days_df = base_group_df.copy()
                effort_days_df = effort_days_df.groupby(['group_id','name'])['date'].nunique().reset_index()
                effort_days_df.rename(columns={'date': 'count'}, inplace=True)
                summary['effort_days'] = effort_days_df.sort_values(by='count', ascending=False).to_dict(orient='records')
                    
                # Active Days
                # Calculate which days each site was active for from a start date to an end date
                if startDate:
                    start_date = datetime.strptime(startDate.split(' ')[0], '%Y-%m-%d')
                else:
                    start_date = base_group_df['timestamp'].min()

                if endDate:
                    end_date = datetime.strptime(endDate.split(' ')[0], '%Y-%m-%d')
                else:
                    end_date = base_group_df['timestamp'].max()

                # Create a new id that will be assigned to sites that have the same site tag and coordinates
                group_active_df = base_group_df.copy()
                group_active_df = group_active_df.drop_duplicates(subset=['id', 'timestamp', 'group_id'])

                # Add a count column to the trap_active_df DataFrame and set it to 1
                group_active_df['count'] = 1

                # Group by 'site_id' and count the unique dates to get the number of days each site captured an image
                group_active_df = group_active_df.groupby(['group_id', 'name', 'date'])['count'].sum().reset_index()
                group_active_dict = group_active_df.sort_values(by='date', ascending=True).to_dict(orient='records')

                # Get the names of the sites for each site_id and convert to a list
                unit_names = group_active_df.groupby('group_id')['name'].unique().explode().tolist()

                max_count = group_active_df['count'].max()
                if np.isnan(max_count):
                    max_count = 0

                summary['active_days'] = {
                    'active_dict': group_active_dict,
                    'start_date': start_date,
                    'end_date': end_date,
                    'unit_names': unit_names,
                    'max_count' : int(max_count)
                }
                
                # Unit Counts            
                group_counts_df = df[['id', 'group_id', 'group_name']].copy()
                group_counts_df = group_counts_df.drop_duplicates(subset=['id', 'group_id'])
                group_counts_df['count'] = 1   

                # rename columns
                group_counts_df.rename(columns={'group_name': 'name'}, inplace=True)

                # Combine sites with same id and sum theur counts
                group_counts_df = group_counts_df.groupby(['group_id', 'name'])['count'].sum().reset_index()
                
                # Convert the DataFrame to a list of dictionaries
                summary['unit_counts'] = group_counts_df.sort_values(by='count', ascending=False).to_dict(orient='records')


            # Species abundance
            if normaliseBySite: # Normalise by site effort (Relative Abundance)
                trapgroups = db.session.query(
                                        Trapgroup.tag, 
                                        Trapgroup.latitude, 
                                        Trapgroup.longitude,
                                        func.count(distinct(func.date(Image.corrected_timestamp)))
                                    )\
                                    .join(Camera, Camera.trapgroup_id==Trapgroup.id)\
                                    .join(Image)\
                                    .outerjoin(Sitegroup, Trapgroup.sitegroups)\
                                    .filter(Trapgroup.survey_id.in_(survey_ids))
                                        

                if sites and sites != '0' and sites != '-1' and groups and groups != '0' and groups != '-1':
                    trapgroups = trapgroups.filter(or_(Trapgroup.id.in_(sites), Sitegroup.id.in_(groups)))
                elif sites and sites != '0' and sites != '-1':
                    trapgroups = trapgroups.filter(Trapgroup.id.in_(sites))
                elif groups and groups != '0' and groups != '-1':
                    trapgroups = trapgroups.filter(Sitegroup.id.in_(groups))

                trapgroups = trapgroups.group_by(Trapgroup.tag, Trapgroup.latitude, Trapgroup.longitude).order_by(Trapgroup.tag).all()

                site_counts = pd.DataFrame(trapgroups, columns=['name', 'latitude', 'longitude', 'count'])
                label_list.append(GLOBALS.unknown_id)
                species_df = df[~df['label_id'].isin(label_list)].copy()
                species_df = species_df[['id','label_id','species', 'name', 'latitude', 'longitude']]
                species_df.drop_duplicates(subset=['id','species', 'name', 'latitude', 'longitude'], inplace=True)	

                species_df['count'] = 1
                species_df = species_df.groupby(['species', 'name', 'latitude', 'longitude'])['count'].sum().reset_index()

                site_counts.rename(columns={'count': 'effort_days'}, inplace=True)

                # create empty new dataframe where we will store the new values
                new_df = pd.DataFrame(columns=['species', 'name', 'latitude', 'longitude', 'count', 'effort_days'])
                for species in species_df['species'].unique():
                    species_df_temp = species_df[species_df['species'] == species]
                    species_df_temp = species_df_temp.merge(site_counts, on=['name', 'latitude', 'longitude'], how='outer')
                    species_df_temp['count'] = species_df_temp['count'].fillna(0)
                    species_df_temp['species'] = species
                    new_df = new_df.append(species_df_temp)

                species_df = new_df.reset_index(drop=True)

                # calculate abundance
                species_df['count'] = species_df['count'].astype(float)
                species_df['effort_days'] = species_df['effort_days'].astype(float)
                species_df['abundance'] = species_df['count']/species_df['effort_days'] * 100
                
                species_df = species_df.groupby('species')['abundance'].mean().reset_index()
                species_df['abundance'] = species_df['abundance'].round(3)
                species_df.rename(columns={'abundance': 'count'}, inplace=True)
                summary['species_count'] = species_df.sort_values(by='count', ascending=False).to_dict(orient='records')

            else: # Total counts (Absolute Abundance)
                summary['species_count'] = species_count.sort_values(by='count', ascending=False).to_dict(orient='records')

        status = 'SUCCESS'
        error = None

    except Exception as exc:
        app.logger.info(' ')
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(traceback.format_exc())
        app.logger.info('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
        app.logger.info(' ')
        status = 'FAILURE'
        error = str(exc)

    finally:
        db.session.remove()

    return { 'status': status, 'error': error, 'summary': summary }

@celery.task(bind=True, ignore_result=True)
def clean_up_R_results(self,R_type, folder, user_name):
    ''' Deletes any R results files for the given user folder and R type '''
    try: 
        # Check if delete file task is already in queue
        scheduled_files = []
        inspector = celery.control.inspect()
        inspector_scheduled = inspector.scheduled()

        for worker in inspector_scheduled:
            for task in inspector_scheduled[worker]:
                if 'deleteFile' in task['request']['name']:
                        scheduled_files.append(task['request']['kwargs']['fileName'])

        if R_type == 'activity':
            isFile = re.compile('^' + user_name + '_Activity_Pattern_', re.I)
        elif R_type == 'occupancy':
            isFile = re.compile('^' + user_name + '_Occupancy_', re.I)
        elif R_type == 'scr':
            isFile = re.compile('^' + user_name + '_SCR_', re.I)
        else:
            isFile = re.compile('^(' + user_name + '_Activity_Pattern_|' + user_name + '_Occupancy_|' + user_name + '_SCR_)', re.I)

        sourceBucket = Config.BUCKET
        s3Folder = folder + '/docs'

        if Config.DEBUGGING: app.logger.info('Cleaning up R results (' + R_type + ') in ' + folder + ' for ' + user_name)

        for dirpath, folders, filenames in s3traverse(sourceBucket, s3Folder):
            files = list(filter(isFile.search, filenames))
            for file in files:
                fileName = s3Folder + '/' + file
                if fileName not in scheduled_files:
                    deleteFile.apply_async(kwargs={'fileName': fileName}, countdown=3600)

    except:
        pass

    finally:
        db.session.remove()

    return True
